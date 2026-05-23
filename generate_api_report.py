# -*- coding: utf-8 -*-
"""
Regenerate MiniGate_API_Report.xlsx with all current API endpoints.
Run: python generate_api_report.py
Requires: openpyxl  (pip install openpyxl)
"""
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Color palette ─────────────────────────────────────────────────────────────
BLUE   = "1D4ED8"
LBLUE  = "EFF6FF"
DGREY  = "374151"
WHITE  = "FFFFFF"
GREEN  = "16A34A"
AMBER  = "D97706"
RED    = "DC2626"
LGREY  = "F9FAFB"
BORDER_COLOR = "E5E7EB"

thin = Side(style="thin", color=BORDER_COLOR)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hfill(color):
    return PatternFill("solid", fgColor=color.lstrip("#"))

def bold(color="000000", size=10):
    return Font(bold=True, color=color.lstrip("#"), size=size)

def normal(color="374151", size=9):
    return Font(bold=False, color=color.lstrip("#"), size=size)

METHOD_COLORS = {"GET": GREEN, "POST": BLUE, "PATCH": AMBER, "DELETE": RED}

def method_font(method):
    return Font(bold=True, color=METHOD_COLORS.get(method, DGREY), size=9)

WRAP   = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")

# ── All API endpoints ─────────────────────────────────────────────────────────
SECTIONS = [

    # ── AUTH ──────────────────────────────────────────────────────────────────
    ("Authentication", [
        ("POST", "/api/accounts/otp/send/",                      "Send OTP to mobile",             "{ mobile }",                                              "200 { success }"),
        ("POST", "/api/accounts/otp/verify/",                    "Verify OTP",                     "{ mobile, otp_code }",                                    "200 { success }"),
        ("POST", "/api/accounts/login/email/",                   "Email + password login",         "{ email, password }",                                     "200 { tokens, data, home_route }"),
        ("POST", "/api/accounts/login/mobile/",                  "Mobile OTP login",               "{ mobile, otp_code }",                                    "200 { tokens, data, home_route, features }"),
        ("POST", "/api/accounts/token/refresh/",                 "Refresh JWT access token",       "{ refresh }",                                             "200 { access }"),
        ("GET",  "/api/accounts/me/",                            "Current user profile",           "Bearer token",                                            "200 { success, data }"),
        ("GET",  "/api/accounts/my-home/",                       "Home screen helper",             "?mobile=",                                                "200 { role, home_route }"),
        ("GET",  "/api/accounts/onboarding/countries/",          "List countries",                 "—",                                                       "200 []"),
        ("GET",  "/api/accounts/onboarding/cities/",             "List cities",                    "—",                                                       "200 []"),
        ("GET",  "/api/accounts/onboarding/societies/",          "Societies (onboarding)",         "—",                                                       "200 []"),
        ("GET",  "/api/accounts/onboarding/buildings/",          "Buildings for society",          "?society=<id>",                                           "200 []"),
        ("GET",  "/api/accounts/onboarding/flats/",              "Flats for society",              "?society=<id>",                                           "200 { count, results }"),
        ("POST", "/api/accounts/onboarding/complete/",           "Complete resident onboarding",   "{ mobile, full_name, society_id, flat_number, … }",       "201 { tokens, data }"),
        ("GET",  "/api/accounts/onboarding/approval-status/",    "Poll resident approval status",  "?mobile=",                                                "200 { status }"),
    ]),

    # ── PLATFORM ADMIN ────────────────────────────────────────────────────────
    ("Platform Admin — Dashboard", [
        ("GET",   "/api/platform-admin/dashboard/stats/",         "Platform-wide KPIs",            "—",                                                      "200 { societies, users, … }"),
        ("GET",   "/api/platform-admin/dashboard/societies/",     "Society list for dashboard",    "?search, ?status, ?plan, ?city, ?ordering, ?page",       "200 paginated"),
    ]),
    ("Platform Admin — Society CRUD", [
        ("GET",   "/api/platform-admin/create-society/societies/",       "List all societies",    "?status, ?plan, ?city, ?search, ?ordering",               "200 paginated"),
        ("POST",  "/api/platform-admin/create-society/societies/",       "Create society",        "{ name, city, plan, total_flats, admin_email }",          "201 { data }"),
        ("GET",   "/api/platform-admin/create-society/societies/{id}/",  "Retrieve society",      "—",                                                       "200 { data }"),
        ("PATCH", "/api/platform-admin/create-society/societies/{id}/",  "Update society",        "{ field: value }",                                        "200 { data }"),
        ("GET",   "/api/platform-admin/society-management/",             "Society management",    "—",                                                       "200 []"),
        ("GET",   "/api/platform-admin/society-admins/",                 "Society admins list",   "—",                                                       "200 paginated"),
    ]),
    ("Platform Admin — Plans, Users & Settings", [
        ("GET",   "/api/platform-admin/subscription-plans/",         "List plans",                "—",                                                       "200 paginated"),
        ("POST",  "/api/platform-admin/subscription-plans/",         "Create plan",               "{ name, slug, monthly_price, annual_price, max_flats }",  "201 { data }"),
        ("GET",   "/api/platform-admin/subscription-plans/{id}/",    "Retrieve plan",             "—",                                                       "200 { data }"),
        ("PATCH", "/api/platform-admin/subscription-plans/{id}/",    "Update plan",               "{ field: value }",                                        "200 { data }"),
        ("DELETE","/api/platform-admin/subscription-plans/{id}/",    "Delete plan",               "—",                                                       "200 { success }"),
        ("GET",   "/api/platform-admin/global-users/",               "All users (platform)",      "?search, ?role",                                          "200 paginated"),
        ("GET",   "/api/platform-admin/global-reports/overview/",    "Global overview",           "—",                                                       "200 { data }"),
        ("GET",   "/api/platform-admin/global-reports/revenue/",     "Revenue report",            "—",                                                       "200 { data }"),
        ("GET",   "/api/platform-admin/global-reports/society-growth/","Society growth",          "—",                                                       "200 { data }"),
        ("GET",   "/api/platform-admin/audit-logs/",                 "Platform audit logs",       "?page",                                                   "200 paginated"),
        ("GET",   "/api/platform-admin/system-settings/",            "System settings GET",       "—",                                                       "200 { data }"),
        ("PATCH", "/api/platform-admin/system-settings/",            "System settings PATCH",     "{ maintenance_mode, … }",                                 "200 { data }"),
    ]),

    # ── ROLES & PERMISSIONS ───────────────────────────────────────────────────
    ("Roles & Permissions", [
        ("GET",  "/api/roles-permissions/roles/",       "List all roles",           "—",              "200 { results }"),
        ("GET",  "/api/roles-permissions/roles/{id}/",  "Retrieve role detail",     "—",              "200 { data }"),
        ("GET",  "/api/roles-permissions/users/",       "List user profiles",       "?role, ?society","200 paginated"),
    ]),

    # ── SOCIETY ADMIN ─────────────────────────────────────────────────────────
    ("Society Admin — Dashboard, Buildings & Flats", [
        ("GET",   "/api/society-admin/dashboard/",           "Society dashboard",         "?society= (optional, auto-detected)", "200 { data }"),
        ("GET",   "/api/society-admin/buildings/",           "List buildings",             "?society=",                          "200 []"),
        ("POST",  "/api/society-admin/buildings/",           "Create building",            "{ name, floors, … }",               "201 { data }"),
        ("GET",   "/api/society-admin/buildings/{id}/",      "Retrieve building",          "—",                                  "200 { data }"),
        ("PATCH", "/api/society-admin/buildings/{id}/",      "Update building",            "{ field: value }",                  "200 { data }"),
        ("DELETE","/api/society-admin/buildings/{id}/",      "Delete building",            "—",                                  "200 { success }"),
        ("GET",   "/api/society-admin/flats/",               "List flats",                 "?society=, ?building=",             "200 paginated"),
        ("POST",  "/api/society-admin/flats/",               "Create flat",                "{ flat_number, building, … }",      "201 { data }"),
        ("PATCH", "/api/society-admin/flats/{id}/",          "Update flat",                "{ field: value }",                  "200 { data }"),
        ("DELETE","/api/society-admin/flats/{id}/",          "Delete flat",                "—",                                  "200 { success }"),
    ]),
    ("Society Admin — Residents & Approvals", [
        ("GET",   "/api/society-admin/residents/",            "Residents list",               "?society=",                      "200 paginated"),
        ("POST",  "/api/society-admin/residents/",            "Add resident manually",        "{ full_name, mobile, flat, … }", "201 { data }"),
        ("GET",   "/api/society-admin/residents/{id}/",       "Resident detail",              "—",                              "200 { data }"),
        ("PATCH", "/api/society-admin/residents/{id}/",       "Update resident",              "{ field: value }",               "200 { data }"),
        ("GET",   "/api/society-admin/approvals/",            "Approval queue",               "?society=",                      "200 []"),
        ("POST",  "/api/society-admin/approvals/",            "Approve / reject resident",    "{ profile_id, action }",         "200 { success }"),
    ]),
    ("Society Admin — Finance", [
        ("GET",   "/api/society-admin/payments/overview/",        "Payment overview",            "?society=",      "200 { data }"),
        ("GET",   "/api/society-admin/fund-dashboard/",           "Fund dashboard",              "?society=",      "200 { data }"),
        ("GET",   "/api/society-admin/maintenance-expenses/",     "Maintenance expenses list",   "?society=",      "200 paginated"),
        ("POST",  "/api/society-admin/maintenance-expenses/",     "Create expense",              "{ title, category, amount, … }", "201 { data }"),
        ("PATCH", "/api/society-admin/maintenance-expenses/{id}/","Update expense",              "{ field: value }","200 { data }"),
        ("DELETE","/api/society-admin/maintenance-expenses/{id}/","Delete expense",              "—",               "200 { success }"),
        ("GET",   "/api/society-admin/monthly-statements/",       "Monthly statements",          "?society=",      "200 paginated"),
        ("GET",   "/api/society-admin/analytics/",                "Analytics",                   "?society=",      "200 { data }"),
    ]),
    ("Society Admin — Misc", [
        ("GET",   "/api/society-admin/notice-board/",          "Notice board list",         "?society=",             "200 paginated"),
        ("POST",  "/api/society-admin/notice-board/",          "Create notice",             "{ title, content, notice_type, … }", "201 { data }"),
        ("GET",   "/api/society-admin/notice-board/{id}/",     "Retrieve notice",           "—",                     "200 { data }"),
        ("PATCH", "/api/society-admin/notice-board/{id}/",     "Update notice",             "{ field: value }",      "200 { data }"),
        ("DELETE","/api/society-admin/notice-board/{id}/",     "Delete notice",             "—",                     "200 { success, message }"),
        ("GET",   "/api/society-admin/complaints/",            "Complaints list",           "?society=",             "200 paginated"),
        ("GET",   "/api/society-admin/staff-guards/",          "Staff & guards list",       "?society=",             "200 paginated"),
        ("POST",  "/api/society-admin/staff-guards/",          "Add staff/guard",           "{ name, role, … }",     "201 { data }"),
        ("GET",   "/api/society-admin/vendors/",               "Vendors list",              "?society=",             "200 paginated"),
        ("GET",   "/api/society-admin/visitors/",              "Visitor log",               "?society=",             "200 paginated"),
        ("GET",   "/api/society-admin/security/",              "Security oversight",        "?society=",             "200 { data }"),
        ("GET",   "/api/society-admin/roles-access/",          "Roles & access",            "?society=",             "200 { data }"),
        ("GET",   "/api/society-admin/settings/",              "Society settings GET",      "?society=",             "200 { data }"),
        ("PATCH", "/api/society-admin/settings/",              "Society settings PATCH",    "{ admin_email, … }",    "200 { data }"),
        ("GET",   "/api/society-admin/notifications/",         "Notifications",             "?society=",             "200 paginated"),
        ("GET",   "/api/society-admin/audit-logs/",            "Audit logs",                "?society=",             "200 paginated"),
    ]),

    # ── RESIDENT ─────────────────────────────────────────────────────────────
    ("Resident — Core", [
        ("GET",  "/api/resident/dashboard/",                           "Resident dashboard",          "?society=, ?flat=",                                 "200 { data }"),
        ("GET",  "/api/resident/payments/",                            "My payments",                 "?flat=",                                            "200 paginated"),
        ("GET",  "/api/resident/notices/",                             "Published notices",           "?society=",                                         "200 paginated"),
        ("GET",  "/api/resident/complaints/",                          "My complaints",               "?flat=",                                            "200 paginated"),
        ("POST", "/api/resident/complaints/",                          "Submit complaint",            "{ title, description, category, flat, society }",   "201 { data }"),
        ("PATCH","/api/resident/complaints/{id}/",                     "Update complaint",            "{ description }",                                   "200 { data }"),
        ("GET",  "/api/resident/visitors/",                            "My visitor log",              "?flat=",                                            "200 paginated"),
        ("GET",  "/api/resident/sos/",                                 "SOS alerts",                  "?flat=",                                            "200 paginated"),
        ("GET",  "/api/resident/maintenance-transparency/",            "Published expenses",          "?society=, ?flat=",                                 "200 { data }"),
        ("GET",  "/api/resident/maintenance-transparency/expenses/{id}/proof/", "Download expense proof", "?society=",                                    "302 redirect / 200 { filename }"),
    ]),
    ("Resident — Monthly Statements", [
        ("GET",  "/api/resident/monthly-statements/",                  "Published statements list",   "?year=",                                            "200 paginated"),
        ("GET",  "/api/resident/monthly-statements/{id}/",             "Statement detail (published)", "—",                                               "200 { data }"),
        ("GET",  "/api/resident/monthly-statements/{id}/download-pdf/","Download statement PDF",      "—",                                                 "200 PDF"),
    ]),
    ("Resident — Profile & Sub-Resources", [
        ("GET",   "/api/resident/profile/my-flats/",                   "My flat links",               "—",                                                  "200 { count, results }"),
        ("POST",  "/api/resident/profile/my-flats/add/",               "Request flat link",           "{ society_id, flat_number }",                       "201 { data }"),
        ("POST",  "/api/resident/profile/my-flats/{id}/switch/",       "Switch primary flat",         "—",                                                  "200 { success }"),
        ("DELETE","/api/resident/profile/my-flats/{id}/remove/",       "Remove flat link",            "—",                                                  "200 { success }"),
        ("GET",   "/api/resident/profile/family/",                     "Family members list",         "—",                                                  "200 paginated"),
        ("POST",  "/api/resident/profile/family/",                     "Add family member",           "{ name, relation, phone, resident, flat }",          "201 { data }"),
        ("GET",   "/api/resident/profile/family/{id}/",                "Retrieve family member",      "—",                                                  "200 { data }"),
        ("PATCH", "/api/resident/profile/family/{id}/",                "Update family member",        "{ phone }",                                          "200 { data }"),
        ("DELETE","/api/resident/profile/family/{id}/",                "Delete family member",        "—",                                                  "200 { success }"),
        ("GET",   "/api/resident/profile/vehicles/",                   "Vehicles list",               "—",                                                  "200 paginated"),
        ("POST",  "/api/resident/profile/vehicles/",                   "Add vehicle",                 "{ vehicle_name, vehicle_type, plate_number, … }",   "201 { data }"),
        ("PATCH", "/api/resident/profile/vehicles/{id}/",              "Update vehicle",              "{ parking_slot }",                                   "200 { data }"),
        ("DELETE","/api/resident/profile/vehicles/{id}/",              "Delete vehicle",              "—",                                                  "200 { success }"),
        ("GET",   "/api/resident/profile/pets/",                       "Pets list",                   "—",                                                  "200 paginated"),
        ("POST",  "/api/resident/profile/pets/",                       "Add pet",                     "{ name, pet_type, gender, color }",                  "201 { data }"),
        ("DELETE","/api/resident/profile/pets/{id}/",                  "Delete pet",                  "—",                                                  "200 { success }"),
        ("GET",   "/api/resident/profile/daily-help/",                 "Daily help list",             "—",                                                  "200 paginated"),
        ("POST",  "/api/resident/profile/daily-help/",                 "Add daily help",              "{ name, help_type, monthly_salary, status }",        "201 { data }"),
        ("DELETE","/api/resident/profile/daily-help/{id}/",            "Delete daily help",           "—",                                                  "200 { success }"),
    ]),

    # ── ACCOUNTANT ────────────────────────────────────────────────────────────
    ("Accountant — Billing Dashboard", [
        ("GET",  "/api/accountant/dashboard/",  "Billing KPIs + 12-month history",  "Bearer accountant",  "200 { collected_this_month, outstanding, defaulters, avg_collection_pct, monthly_history }"),
    ]),
    ("Accountant — Payment Collection: Dues", [
        ("GET",   "/api/accountant/payment-collection/dues/",                    "List dues",              "?month=YYYY-MM, ?status, ?building, ?flat, ?page",         "200 paginated"),
        ("POST",  "/api/accountant/payment-collection/dues/",                    "Create due manually",    "{ flat, month, amount, due_date, description }",           "201 { data }"),
        ("GET",   "/api/accountant/payment-collection/dues/{id}/",               "Retrieve due",           "—",                                                        "200 { data }"),
        ("PATCH", "/api/accountant/payment-collection/dues/{id}/",               "Update unpaid due",      "{ amount, due_date, description }",                        "200 { data }"),
        ("DELETE","/api/accountant/payment-collection/dues/{id}/",               "Delete unpaid due",      "—",                                                        "200 { success }"),
        ("POST",  "/api/accountant/payment-collection/dues/generate/",           "Bulk-generate dues",     "{ year, month, amount, due_day?, description? }",          "201 { created, skipped }"),
        ("POST",  "/api/accountant/payment-collection/dues/{id}/mark-paid/",     "Mark due as paid",       "{ payment_method, payment_date?, description? }",          "200 { due, payment_id }"),
    ]),
    ("Accountant — Payment Collection: Payments", [
        ("GET",   "/api/accountant/payment-collection/payments/",                "List payment records",   "?month, ?payment_type, ?payment_method, ?flat, ?resident", "200 paginated"),
        ("POST",  "/api/accountant/payment-collection/payments/",                "Record manual payment",  "{ flat, resident, payment_type, payment_method, amount, payment_date? }", "201 { data }"),
        ("GET",   "/api/accountant/payment-collection/payments/{id}/",           "Retrieve payment",       "—",                                                        "200 { data }"),
        ("PATCH", "/api/accountant/payment-collection/payments/{id}/",           "Update payment",         "{ description?, payment_date?, payment_method? }",         "200 { data }"),
        ("DELETE","/api/accountant/payment-collection/payments/{id}/",           "Delete payment",         "—",                                                        "200 { success }"),
    ]),
    ("Accountant — Pending Dues", [
        ("GET",  "/api/accountant/payment-collection/pending-dues/",                   "Pending dues + KPI summary",  "?search, ?status, ?building, ?month, ?ordering",  "200 { summary, count, results }"),
        ("GET",  "/api/accountant/payment-collection/pending-dues/summary/",           "KPI cards only",               "—",                                              "200 { data }"),
        ("POST", "/api/accountant/payment-collection/pending-dues/{id}/mark-paid/",    "Mark pending due paid",        "{ payment_method, payment_date?, description? }", "200 { due, payment_id }"),
        ("POST", "/api/accountant/payment-collection/pending-dues/send-reminders/",    "Queue reminders",              "{ status? }",                                    "200 { success, recipients }"),
    ]),
    ("Accountant — Track Payments", [
        ("GET",  "/api/accountant/track-payments/",         "Paginated payment list",   "?search, ?month, ?status, ?payment_method, ?payment_type, ?building, ?flat, ?ordering", "200 paginated"),
        ("GET",  "/api/accountant/track-payments/summary/", "KPI aggregate cards",      "same filters as list",                                                                   "200 { data }"),
        ("GET",  "/api/accountant/track-payments/export/",  "CSV download",             "same filters as list",                                                                   "200 CSV"),
        ("GET",  "/api/accountant/track-payments/{id}/",    "Single payment detail",    "—",                                                                                      "200 { data }"),
    ]),
    ("Accountant — Fund Dashboard", [
        ("GET",  "/api/accountant/fund-dashboard/",  "Fund KPIs + latest expenses + monthly trend",  "?months=12",  "200 { kpi, latest_expenses, monthly_trend }"),
    ]),
    ("Accountant — Maintenance Expenses", [
        ("GET",   "/api/accountant/maintenance-expenses/",               "List expenses",        "?category, ?is_published, ?search, ?year, ?month, ?ordering",               "200 paginated"),
        ("POST",  "/api/accountant/maintenance-expenses/",               "Create expense",       "{ title, category, amount, vendor_name, payment_mode, invoice_number, building_area, expense_date, is_published, notes }", "201 { data }"),
        ("GET",   "/api/accountant/maintenance-expenses/summary/",       "Category breakdown",   "?year, ?month",                                                              "200 { data }"),
        ("GET",   "/api/accountant/maintenance-expenses/{id}/",          "Retrieve expense",     "—",                                                                          "200 { data }"),
        ("PATCH", "/api/accountant/maintenance-expenses/{id}/",          "Update expense",       "{ field: value }",                                                           "200 { data }"),
        ("DELETE","/api/accountant/maintenance-expenses/{id}/",          "Delete expense",       "—",                                                                          "200 { success }"),
        ("POST",  "/api/accountant/maintenance-expenses/{id}/publish/",  "Publish expense",      "—",                                                                          "200 { data, is_published: true }"),
        ("POST",  "/api/accountant/maintenance-expenses/{id}/unpublish/","Retract to draft",     "—",                                                                          "200 { data, is_published: false }"),
        ("GET",   "/api/accountant/maintenance-expenses/{id}/proof/",    "Download proof",       "—",                                                                          "302 redirect or 200 { filename }"),
    ]),
    ("Accountant — Monthly Statements", [
        ("GET",   "/api/accountant/monthly-statements/",                     "List statements",       "?is_published, ?year, ?page",              "200 paginated"),
        ("POST",  "/api/accountant/monthly-statements/generate/",            "Generate statement",    "{ year, month, opening_balance?, notes? }", "201 { data }"),
        ("GET",   "/api/accountant/monthly-statements/{id}/",                "Statement detail",      "—",                                        "200 { data }"),
        ("POST",  "/api/accountant/monthly-statements/{id}/publish/",        "Publish statement",     "—",                                        "200 { data }"),
        ("POST",  "/api/accountant/monthly-statements/{id}/unpublish/",      "Retract statement",     "—",                                        "200 { data }"),
        ("POST",  "/api/accountant/monthly-statements/{id}/upload-proof/",   "Upload proof files",    "multipart, key=files (max 10)",            "201 { data }"),
        ("DELETE","/api/accountant/monthly-statements/{id}/delete-proof/",   "Delete proof file",     "?doc_id=<id>",                             "200 { success }"),
        ("GET",   "/api/accountant/monthly-statements/{id}/download-pdf/",   "Download PDF",          "—",                                        "200 PDF"),
        ("GET",   "/api/accountant/monthly-statements/{id}/export-excel/",   "Download Excel (.xlsx)","—",                                        "200 XLSX"),
    ]),
    ("Accountant — Generate Receipts", [
        ("GET",  "/api/accountant/generate-receipts/",             "Receipts list",                "?month=YYYY-MM, ?flat, ?resident, ?payment_type, ?search", "200 paginated"),
        ("GET",  "/api/accountant/generate-receipts/bulk-pdf/",    "All receipts single PDF",      "same filters as list",                                     "200 PDF (A4 multi-page)"),
        ("GET",  "/api/accountant/generate-receipts/bulk-csv/",    "All receipts CSV",             "same filters as list",                                     "200 CSV"),
        ("GET",  "/api/accountant/generate-receipts/{id}/",        "JSON receipt detail",          "—",                                                        "200 { data }"),
        ("GET",  "/api/accountant/generate-receipts/{id}/pdf/",    "Single receipt PDF",           "—",                                                        "200 PDF (A5, ReportLab branded)"),
    ]),
    ("Accountant — Payment Reports", [
        ("GET",  "/api/accountant/payment-reports/",               "Analytics JSON report",        "?months=12, ?year, ?payment_type, ?payment_method",       "200 { data: { total_payments, by_method, by_type, monthly_trend } }"),
        ("GET",  "/api/accountant/payment-reports/download-pdf/",  "Analytics PDF download",       "same filters as JSON report",                              "200 PDF"),
    ]),
    ("Accountant — Export Reports", [
        ("GET",  "/api/accountant/export-reports/payments/",         "Payments CSV",               "?month, ?year, ?building, ?flat",                         "200 CSV"),
        ("GET",  "/api/accountant/export-reports/payments/pdf/",     "Payments PDF table",         "same filters",                                            "200 PDF"),
        ("GET",  "/api/accountant/export-reports/payments/tally/",   "Payments TallyPrime XML",    "?month, ?year, ?building, ?flat",                         "200 XML (TallyPrime ENVELOPE/VOUCHER format)"),
        ("GET",  "/api/accountant/export-reports/dues/",             "Dues CSV",                   "?month, ?year, ?status, ?building, ?flat",               "200 CSV"),
        ("GET",  "/api/accountant/export-reports/dues/pdf/",         "Dues PDF table",             "same filters",                                            "200 PDF"),
        ("GET",  "/api/accountant/export-reports/expenses/",         "Expenses CSV",               "?month, ?year, ?is_published",                            "200 CSV"),
        ("GET",  "/api/accountant/export-reports/expenses/pdf/",     "Expenses PDF table",         "same filters",                                            "200 PDF"),
        ("GET",  "/api/accountant/export-reports/statements/",       "Statements CSV",             "?year, ?is_published",                                    "200 CSV"),
        ("GET",  "/api/accountant/export-reports/statements/pdf/",   "Statements PDF table",       "same filters",                                            "200 PDF"),
    ]),

    # ── SECURITY GUARD ────────────────────────────────────────────────────────
    ("Security Guard", [
        ("GET",  "/api/security-guard/dashboard/",         "Guard dashboard",            "Bearer guard-token",    "200 { data }"),
        ("GET",  "/api/security-guard/gate-entry/",        "Gate entry log",             "?date, ?search",        "200 paginated"),
        ("POST", "/api/security-guard/gate-entry/",        "Log gate entry",             "{ visitor_name, flat, purpose }",  "201 { data }"),
        ("GET",  "/api/security-guard/visitor-log/",       "Visitor history",            "?date, ?flat",          "200 paginated"),
        ("GET",  "/api/security-guard/vehicle-tracking/",  "Vehicle entries",            "?plate",                "200 paginated"),
        ("GET",  "/api/security-guard/emergency-alerts/",  "Emergency alerts",           "—",                     "200 paginated"),
        ("GET",  "/api/security-guard/shift-management/",  "Shift schedule",             "—",                     "200 { data }"),
    ]),

    # ── MAINTENANCE STAFF ─────────────────────────────────────────────────────
    ("Maintenance Staff", [
        ("GET",  "/api/maintenance-staff/dashboard/",          "Staff dashboard",         "—",  "200 { data }"),
        ("GET",  "/api/maintenance-staff/assigned-tasks/",     "Assigned tasks",          "—",  "200 paginated"),
        ("PATCH","/api/maintenance-staff/task-updates/",       "Update task status",      "{ task_id, status }",  "200 { data }"),
        ("GET",  "/api/maintenance-staff/work-history/",       "Work history",            "—",  "200 paginated"),
        ("GET",  "/api/maintenance-staff/materials-request/",  "Materials requests",      "—",  "200 paginated"),
        ("GET",  "/api/maintenance-staff/schedule/",           "Schedule",                "—",  "200 { data }"),
    ]),

    # ── SUPPORT STAFF ─────────────────────────────────────────────────────────
    ("Support Staff", [
        ("GET",  "/api/support-staff/dashboard/",           "Support dashboard",     "—",  "200 { data }"),
        ("GET",  "/api/support-staff/assigned-tickets/",    "Assigned tickets",      "—",  "200 paginated"),
        ("PATCH","/api/support-staff/ticket-updates/",      "Update ticket",         "{ ticket_id, status }",  "200 { data }"),
        ("GET",  "/api/support-staff/escalations/",         "Escalations",           "—",  "200 paginated"),
        ("GET",  "/api/support-staff/service-history/",     "Service history",       "—",  "200 paginated"),
    ]),

    # ── DELIVERY PARTNER ──────────────────────────────────────────────────────
    ("Delivery Partner", [
        ("GET",  "/api/delivery-partner/dashboard/",           "Delivery dashboard",     "—",        "200 { data }"),
        ("GET",  "/api/delivery-partner/delivery-requests/",   "Delivery requests",      "—",        "200 paginated"),
        ("POST", "/api/delivery-partner/otp-verification/",    "OTP verification",       "{ otp }",  "200 { success }"),
        ("GET",  "/api/delivery-partner/delivery-history/",    "Delivery history",       "—",        "200 paginated"),
        ("GET",  "/api/delivery-partner/profile/",             "Delivery profile",       "—",        "200 { data }"),
    ]),

    # ── GUEST USER ────────────────────────────────────────────────────────────
    ("Guest User", [
        ("GET",  "/api/guest/dashboard/",       "Guest dashboard",        "—",                          "200 { data }"),
        ("POST", "/api/guest/visit-request/",   "Submit visit request",   "{ host_flat, purpose, … }",  "201 { data }"),
        ("GET",  "/api/guest/host-info/",       "Host information",       "?flat=",                     "200 { data }"),
        ("GET",  "/api/guest/access-pass/",     "Access pass",            "—",                          "200 { data }"),
        ("GET",  "/api/guest/profile/",         "Guest profile",          "—",                          "200 { data }"),
    ]),
]


def build_workbook():
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "API Reference"

    # Column widths
    for i, w in enumerate([9, 58, 36, 60, 32], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title row
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value          = "MiniGate Backend — Complete API Reference"
    c.font           = Font(bold=True, size=14, color=WHITE)
    c.fill           = hfill(BLUE)
    c.alignment      = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column headers
    for col, h in enumerate(["Method", "Endpoint", "Description", "Request (body / params)", "Response"], 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = bold(WHITE, 10)
        c.fill      = hfill(DGREY)
        c.alignment = CENTER if col == 1 else WRAP
        c.border    = BORDER
    ws.row_dimensions[2].height = 18

    row   = 3
    total = 0
    for section_title, endpoints in SECTIONS:
        # Section separator
        ws.merge_cells(f"A{row}:E{row}")
        c = ws[f"A{row}"]
        c.value          = f"  {section_title}"
        c.font           = bold(BLUE, 10)
        c.fill           = hfill(LBLUE)
        c.alignment      = Alignment(horizontal="left", vertical="center")
        c.border         = BORDER
        ws.row_dimensions[row].height = 16
        row += 1

        for i, (method, path, desc, body, resp) in enumerate(endpoints):
            bg = LGREY if i % 2 else WHITE
            for col, val in enumerate([method, path, desc, body, resp], 1):
                c           = ws.cell(row=row, column=col, value=val)
                c.fill      = hfill(bg)
                c.border    = BORDER
                if col == 1:
                    c.font      = method_font(method)
                    c.alignment = CENTER
                else:
                    c.font      = normal()
                    c.alignment = WRAP
            ws.row_dimensions[row].height = 15
            row  += 1
            total += 1

    ws.freeze_panes = "A3"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 12

    for col, h in enumerate(["Module / Section", "# Endpoints"], 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = bold(WHITE)
        c.fill = hfill(BLUE)
        c.alignment = CENTER
        c.border = BORDER

    for i, (name, eps) in enumerate(SECTIONS, 2):
        ws2.cell(row=i, column=1, value=name).border = BORDER
        c = ws2.cell(row=i, column=2, value=len(eps))
        c.border    = BORDER
        c.alignment = CENTER

    last = len(SECTIONS) + 2
    c = ws2.cell(row=last, column=1, value="TOTAL")
    c.font = bold()
    c.border = BORDER
    c = ws2.cell(row=last, column=2, value=total)
    c.font      = bold()
    c.border    = BORDER
    c.alignment = CENTER

    return wb, total


if __name__ == "__main__":
    wb, n = build_workbook()
    path  = "MiniGate_API_Report.xlsx"
    wb.save(path)
    print(f"Saved {path}  —  {n} endpoints across {len(SECTIONS)} sections")
