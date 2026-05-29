import datetime
import io
import logging
from decimal import Decimal

from django.db.models import Q, Sum
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.viewsets import ViewSet
from apps.common.permissions import IsSocietyAdmin
from apps.platform_admin.create_society.models import Society
from apps.resident.payments.models import MaintenanceDue
from apps.society_admin.maintenance_expenses.models import MaintenanceExpense

from .models import MonthlyStatement, StatementProofDocument
from .serializers import (
    GenerateStatementSerializer,
    MonthlyStatementSerializer,
    ProofDocumentSerializer,
)

logger = logging.getLogger(__name__)


def _admin_society(request):
    try:
        sid = request.user.profile.society_id
        if not sid:
            raise ValueError
        return Society.objects.get(pk=sid)
    except Exception:
        from rest_framework.exceptions import PermissionDenied
        raise PermissionDenied("Your account is not linked to any society.")


class _StatementPagination(PageNumberPagination):
    page_size             = 12
    page_size_query_param = "page_size"
    max_page_size         = 60


# ── Internal helpers ───────────────────────────────────────────────────────────

def _compute_live(society_id: int, year: int, month: int) -> dict:
    """Compute financials from live DB for a given society-month."""
    total_collected = (
        MaintenanceDue.objects
        .filter(
            society_id=society_id,
            month__year=year,
            month__month=month,
            status=MaintenanceDue.Status.PAID,
        )
        .aggregate(total=Sum("amount"))["total"]
    ) or Decimal("0")

    total_expenses = (
        MaintenanceExpense.objects
        .filter(
            society_id=society_id,
            expense_date__year=year,
            expense_date__month=month,
        )
        .aggregate(total=Sum("amount"))["total"]
    ) or Decimal("0")

    proof_docs = list(
        MaintenanceExpense.objects
        .filter(
            society_id=society_id,
            expense_date__year=year,
            expense_date__month=month,
            is_published=True,
        )
        .exclude(proof_url="")
        .values_list("proof_url", flat=True)
    )

    return {
        "total_collected": total_collected,
        "total_expenses":  total_expenses,
        "proof_documents": proof_docs,
    }


def _prev_closing_balance(society_id: int, month_start: datetime.date) -> Decimal:
    prev = (
        MonthlyStatement.objects
        .filter(society_id=society_id, month__lt=month_start)
        .order_by("-month")
        .first()
    )
    return prev.closing_balance if prev else Decimal("0")


def _make_summary(opening: Decimal, collected: Decimal, expenses: Decimal) -> str:
    closing = opening + collected - expenses
    return (
        f"Total collected Rs.{collected:,.0f}, "
        f"used Rs.{expenses:,.0f}, "
        f"balance Rs.{closing:,.0f}."
    )


# ── PDF generator (reportlab) ─────────────────────────────────────────────────

def _generate_pdf(stmt: MonthlyStatement) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm,
    )

    styles = getSampleStyleSheet()
    heading  = ParagraphStyle("heading",  parent=styles["Heading1"], fontSize=16, spaceAfter=4)
    subhead  = ParagraphStyle("subhead",  parent=styles["Normal"],   fontSize=10, textColor=colors.grey, spaceAfter=12)
    section  = ParagraphStyle("section",  parent=styles["Heading2"], fontSize=12, spaceBefore=14, spaceAfter=6)
    body     = ParagraphStyle("body",     parent=styles["Normal"],   fontSize=10, spaceAfter=4)
    bold     = ParagraphStyle("bold",     parent=styles["Normal"],   fontSize=10, fontName="Helvetica-Bold")
    summary_style = ParagraphStyle("summary", parent=styles["Normal"], fontSize=10,
                                   backColor=colors.HexColor("#F0F7FF"),
                                   borderPad=8, leading=14)

    story = []

    # ── Header ────────────────────────────────────────────────────────────────
    story.append(Paragraph(stmt.title, heading))
    story.append(Paragraph(f"Society: {stmt.society.name}", subhead))
    if stmt.is_published and stmt.published_at:
        story.append(Paragraph(f"Published on {stmt.published_at.strftime('%Y-%m-%d')}", subhead))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E5E7EB"), spaceAfter=12))

    # ── Financial KPI table ───────────────────────────────────────────────────
    story.append(Paragraph("Financial Summary", section))

    fin_data = [
        ["Opening Balance",  f"Rs. {stmt.opening_balance:,.2f}"],
        ["Total Collected",  f"Rs. {stmt.total_collected:,.2f}"],
        ["Total Expenses",   f"Rs. {stmt.total_expenses:,.2f}"],
        ["Remaining Balance",f"Rs. {stmt.closing_balance:,.2f}"],
    ]
    fin_table = Table(fin_data, colWidths=[80*mm, 80*mm])
    fin_table.setStyle(TableStyle([
        ("FONTNAME",    (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 0), (-1, -1), 11),
        ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#F9FAFB")),
        ("BACKGROUND",  (0, -1), (-1, -1), colors.HexColor("#EFF6FF")),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(fin_table)
    story.append(Spacer(1, 8*mm))

    # ── Expense breakdown ─────────────────────────────────────────────────────
    year  = stmt.month.year
    month = stmt.month.month
    expenses = list(
        MaintenanceExpense.objects
        .filter(society_id=stmt.society_id, expense_date__year=year, expense_date__month=month)
        .order_by("expense_date")
    )

    if expenses:
        story.append(Paragraph("Expense Details", section))
        exp_data = [["Expense", "Category", "Amount", "Date", "Published"]]
        for e in expenses:
            exp_data.append([
                e.title,
                e.get_category_display(),
                f"Rs. {e.amount:,.0f}",
                e.expense_date.strftime("%d %b %Y"),
                "Yes" if e.is_published else "No",
            ])
        col_widths = [55*mm, 38*mm, 30*mm, 28*mm, 20*mm]
        exp_table  = Table(exp_data, colWidths=col_widths)
        exp_table.setStyle(TableStyle([
            ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 9),
            ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
            ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",   (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
            ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ]))
        story.append(exp_table)
        story.append(Spacer(1, 6*mm))

    # ── Proof documents ───────────────────────────────────────────────────────
    all_proofs = list(stmt.proof_documents or [])
    uploaded   = list(stmt.uploaded_proofs.values_list("original_name", flat=True))
    all_proofs += [p for p in uploaded if p not in all_proofs]

    if all_proofs:
        story.append(Paragraph("Proof Documents", section))
        for proof_name in all_proofs:
            story.append(Paragraph(f"• {proof_name}", body))
        story.append(Spacer(1, 6*mm))

    # ── Summary footer ────────────────────────────────────────────────────────
    if stmt.summary:
        story.append(Paragraph(f"Summary: {stmt.summary}", summary_style))

    story.append(Spacer(1, 8*mm))
    story.append(Paragraph(
        f"Generated on {timezone.now().strftime('%d %b %Y %H:%M')} UTC",
        ParagraphStyle("footer", parent=styles["Normal"], fontSize=8, textColor=colors.grey),
    ))

    doc.build(story)
    return buffer.getvalue()


# ── Excel generator (openpyxl) ────────────────────────────────────────────────

def _generate_excel(stmt: MonthlyStatement) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = stmt.month.strftime("%b %Y")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    BLUE   = PatternFill("solid", fgColor="1D4ED8")
    LBLUE  = PatternFill("solid", fgColor="EFF6FF")
    LGREY  = PatternFill("solid", fgColor="F9FAFB")
    W_FONT = Font(bold=True, color="FFFFFF", size=12)
    B_FONT = Font(bold=True, size=11)
    BORDER = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    row = 1

    # Title
    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = stmt.title
    ws[f"A{row}"].font = Font(bold=True, size=14)
    ws[f"A{row}"].fill = BLUE
    ws[f"A{row}"].font = W_FONT
    ws[f"A{row}"].alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    ws[f"A{row}"] = f"Society: {stmt.society.name}"
    ws[f"A{row}"].font = Font(italic=True, size=10)
    row += 1

    if stmt.published_at:
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = f"Published: {stmt.published_at.strftime('%Y-%m-%d')}"
        ws[f"A{row}"].font = Font(italic=True, size=10, color="6B7280")
    row += 2

    # Financial summary
    financials = [
        ("Opening Balance",   stmt.opening_balance),
        ("Total Collected",   stmt.total_collected),
        ("Total Expenses",    stmt.total_expenses),
        ("Remaining Balance", stmt.closing_balance),
    ]
    for i, (label, value) in enumerate(financials):
        is_last = i == len(financials) - 1
        fill = LBLUE if is_last else (LGREY if i % 2 else PatternFill())
        ws[f"A{row}"] = label
        ws[f"B{row}"] = float(value)
        ws[f"A{row}"].font = B_FONT if is_last else Font(size=11)
        ws[f"B{row}"].font = B_FONT if is_last else Font(size=11)
        ws[f"B{row}"].number_format = '#,##0.00'
        ws[f"A{row}"].fill = fill
        ws[f"B{row}"].fill = fill
        ws[f"A{row}"].border = BORDER
        ws[f"B{row}"].border = BORDER
        row += 1

    row += 1

    # Expense table header
    year  = stmt.month.year
    month = stmt.month.month
    expenses = list(
        MaintenanceExpense.objects
        .filter(society_id=stmt.society_id, expense_date__year=year, expense_date__month=month)
        .order_by("expense_date")
    )

    if expenses:
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 12

        headers = ["Expense", "Category", "Amount (Rs.)", "Date", "Published"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = W_FONT
            cell.fill = BLUE
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

        for e in expenses:
            row_data = [
                e.title,
                e.get_category_display(),
                float(e.amount),
                e.expense_date.strftime("%Y-%m-%d"),
                "Yes" if e.is_published else "No",
            ]
            fill = LGREY if row % 2 else PatternFill()
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = fill
                cell.border = BORDER
                if col == 3:
                    cell.number_format = '#,##0.00'
            row += 1

        row += 1

    # Summary
    if stmt.summary:
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = stmt.summary
        ws[f"A{row}"].font = Font(italic=True, size=10)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── ViewSet ────────────────────────────────────────────────────────────────────

class MonthlyStatementViewSet(ViewSet):
    permission_classes = [IsSocietyAdmin]
    """
    ┌────────────────────────────────────────────────────────────────────────────┐
    │ GET    /                    List statements  ?society=<id>                 │
    │ GET    /<id>/               Statement detail (includes uploaded_proofs)    │
    │ POST   /generate/           Generate / re-generate draft statement         │
    │ POST   /<id>/publish/       Publish statement (visible to residents)       │
    │ POST   /<id>/unpublish/     Retract publication                            │
    │ POST   /<id>/upload-proof/  Upload 1-10 PDF/image proof files             │
    │ DELETE /<id>/delete-proof/  Remove a proof document  ?doc_id=<id>         │
    │ GET    /<id>/download-pdf/  Download statement as PDF                      │
    │ GET    /<id>/export-excel/  Download statement as Excel (.xlsx)            │
    └────────────────────────────────────────────────────────────────────────────┘
    """

    pagination_class = _StatementPagination

    # ── List ──────────────────────────────────────────────────────────────────

    def list(self, request):
        society = _admin_society(request)
        qs = (
            MonthlyStatement.objects
            .filter(society=society)
            .prefetch_related("uploaded_proofs")
            .select_related("society")
            .order_by("-month")
        )
        paginator = self.pagination_class()
        page  = paginator.paginate_queryset(qs, request, view=self)
        items = page if page is not None else qs
        data  = MonthlyStatementSerializer(items, many=True).data
        if page is not None:
            return paginator.get_paginated_response(data)
        return Response({"success": True, "count": len(data), "results": data})

    # ── Retrieve ──────────────────────────────────────────────────────────────

    def retrieve(self, request, pk=None):
        try:
            stmt = (
                MonthlyStatement.objects
                .prefetch_related("uploaded_proofs")
                .select_related("society")
                .get(pk=pk)
            )
        except MonthlyStatement.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)
        return Response({"success": True, "data": MonthlyStatementSerializer(stmt).data})

    # ── Generate / Re-generate ────────────────────────────────────────────────

    @action(detail=False, methods=["post"])
    def generate(self, request):
        """
        POST /generate/
        {
          "society": 11, "year": 2026, "month": 5,
          "opening_balance": null,    // null = auto from prev month
          "total_collected": null,    // null = computed from MaintenanceDue
          "total_expenses":  null,    // null = computed from MaintenanceExpense
          "notes": ""
        }
        """
        ser = GenerateStatementSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        d = ser.validated_data

        society    = _admin_society(request)
        society_id = society.pk
        year       = d["year"]
        month      = d["month"]
        month_start = datetime.date(year, month, 1)

        # Compute live values
        live = _compute_live(society_id, year, month)

        # Apply admin overrides if supplied
        total_collected = (
            d["total_collected"]
            if d.get("total_collected") is not None
            else live["total_collected"]
        )
        total_expenses = (
            d["total_expenses"]
            if d.get("total_expenses") is not None
            else live["total_expenses"]
        )

        # Opening balance: use provided value or auto-fetch from prev month
        if d.get("opening_balance") is not None:
            opening_balance = d["opening_balance"]
        else:
            opening_balance = _prev_closing_balance(society_id, month_start)

        closing_balance = opening_balance + total_collected - total_expenses
        summary         = _make_summary(opening_balance, total_collected, total_expenses)

        stmt, created = MonthlyStatement.objects.update_or_create(
            society_id=society_id,
            month=month_start,
            defaults={
                "opening_balance":  opening_balance,
                "total_collected":  total_collected,
                "total_expenses":   total_expenses,
                "closing_balance":  closing_balance,
                "proof_documents":  live["proof_documents"],
                "summary":          summary,
                "notes":            d["notes"],
                "generated_by":     request.user,
            },
        )

        verb = "generated" if created else "regenerated"
        logger.info("STMT_%s | society=%s month=%s-%02d collected=%.0f expenses=%.0f by=%s",
                    verb.upper(), society_id, year, month,
                    float(total_collected), float(total_expenses), request.user)

        stmt.refresh_from_db()
        stmt_data = MonthlyStatementSerializer(
            MonthlyStatement.objects.prefetch_related("uploaded_proofs").get(pk=stmt.pk)
        ).data

        return Response(
            {"success": True, "message": f"Statement {verb} for {stmt.month_label}.", "data": stmt_data},
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )

    # ── Publish ───────────────────────────────────────────────────────────────

    @action(detail=True, methods=["post"])
    def publish(self, request, pk=None):
        """POST /<id>/publish/"""
        try:
            stmt = MonthlyStatement.objects.get(pk=pk)
        except MonthlyStatement.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)

        if stmt.is_published:
            return Response({"detail": "Already published."}, status=400)

        stmt.is_published = True
        stmt.published_at = timezone.now()
        stmt.save(update_fields=["is_published", "published_at", "updated_at"])

        logger.info("STMT_PUBLISH | id=%s month=%s by=%s", stmt.pk, stmt.month, request.user)

        stmt_data = MonthlyStatementSerializer(
            MonthlyStatement.objects.prefetch_related("uploaded_proofs").get(pk=stmt.pk)
        ).data
        return Response({"success": True, "message": f"Statement for {stmt.month_label} published.", "data": stmt_data})

    # ── Unpublish ─────────────────────────────────────────────────────────────

    @action(detail=True, methods=["post"])
    def unpublish(self, request, pk=None):
        """POST /<id>/unpublish/"""
        try:
            stmt = MonthlyStatement.objects.get(pk=pk)
        except MonthlyStatement.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)

        if not stmt.is_published:
            return Response({"detail": "Not published."}, status=400)

        stmt.is_published = False
        stmt.published_at = None
        stmt.save(update_fields=["is_published", "published_at", "updated_at"])

        return Response({"success": True, "message": f"Statement for {stmt.month_label} unpublished."})

    # ── Upload proof documents ────────────────────────────────────────────────

    @action(detail=True, methods=["post"], url_path="upload-proof", parser_classes=[MultiPartParser])
    def upload_proof(self, request, pk=None):
        """
        POST /<id>/upload-proof/
        Content-Type: multipart/form-data
        files: one or more files (key = "files")

        Upload 1-10 proof PDFs/images at once.
        Returns the updated statement with all uploaded_proofs.
        """
        try:
            stmt = MonthlyStatement.objects.get(pk=pk)
        except MonthlyStatement.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)

        files = request.FILES.getlist("files")
        if not files:
            return Response({"detail": "No files provided. Use key 'files' in form-data."}, status=400)
        if len(files) > 10:
            return Response({"detail": "Maximum 10 files per upload."}, status=400)

        created_docs = []
        for f in files:
            doc = StatementProofDocument.objects.create(
                statement=stmt,
                file=f,
                original_name=f.name,
                file_size=f.size,
                uploaded_by=request.user,
            )
            created_docs.append(doc)

        logger.info("STMT_PROOF_UPLOAD | stmt=%s files=%d by=%s",
                    stmt.pk, len(created_docs), request.user)

        stmt_data = MonthlyStatementSerializer(
            MonthlyStatement.objects.prefetch_related("uploaded_proofs").get(pk=stmt.pk)
        ).data
        return Response({
            "success": True,
            "message": f"{len(created_docs)} file(s) uploaded.",
            "uploaded": ProofDocumentSerializer(created_docs, many=True).data,
            "data":    stmt_data,
        }, status=status.HTTP_201_CREATED)

    # ── Delete proof document ─────────────────────────────────────────────────

    @action(detail=True, methods=["delete"], url_path="delete-proof")
    def delete_proof(self, request, pk=None):
        """
        DELETE /<id>/delete-proof/?doc_id=<doc_id>
        Remove a single uploaded proof document.
        """
        try:
            stmt = MonthlyStatement.objects.get(pk=pk)
        except MonthlyStatement.DoesNotExist:
            return Response({"detail": "Statement not found."}, status=404)

        doc_id = request.query_params.get("doc_id")
        if not doc_id:
            return Response({"detail": "doc_id query param required."}, status=400)

        try:
            doc = StatementProofDocument.objects.get(pk=doc_id, statement=stmt)
        except StatementProofDocument.DoesNotExist:
            return Response({"detail": "Proof document not found."}, status=404)

        doc.file.delete(save=False)
        doc.delete()

        logger.info("STMT_PROOF_DELETE | stmt=%s doc=%s by=%s", stmt.pk, doc_id, request.user)
        return Response({"success": True, "message": "Proof document removed."})

    # ── Download PDF ──────────────────────────────────────────────────────────

    @action(detail=True, methods=["get"], url_path="download-pdf")
    def download_pdf(self, request, pk=None):
        """
        GET /<id>/download-pdf/
        Returns a professionally formatted PDF statement.
        """
        try:
            stmt = (
                MonthlyStatement.objects
                .prefetch_related("uploaded_proofs")
                .select_related("society")
                .get(pk=pk)
            )
        except MonthlyStatement.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)

        pdf_bytes = _generate_pdf(stmt)
        filename  = (
            f"{stmt.society.name.replace(' ', '_')}_{stmt.month.strftime('%Y_%m')}_statement.pdf"
        )
        response  = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        logger.info("STMT_PDF_DOWNLOAD | id=%s by=%s", stmt.pk, request.user)
        return response

    # ── Export Excel ──────────────────────────────────────────────────────────

    @action(detail=True, methods=["get"], url_path="export-excel")
    def export_excel(self, request, pk=None):
        """
        GET /<id>/export-excel/
        Returns a formatted .xlsx file.
        """
        try:
            stmt = (
                MonthlyStatement.objects
                .prefetch_related("uploaded_proofs")
                .select_related("society")
                .get(pk=pk)
            )
        except MonthlyStatement.DoesNotExist:
            return Response({"detail": "Not found."}, status=404)

        xlsx_bytes = _generate_excel(stmt)
        filename   = (
            f"{stmt.society.name.replace(' ', '_')}_{stmt.month.strftime('%Y_%m')}_statement.xlsx"
        )
        response = HttpResponse(
            xlsx_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        logger.info("STMT_EXCEL_EXPORT | id=%s by=%s", stmt.pk, request.user)
        return response