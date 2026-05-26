"""
Management command: generate_api_excel
Generates MiniGate_API_Report.xlsx — full API documentation with JWT auth.

Usage:
    python manage.py generate_api_excel --settings=config.settings.development
"""

import datetime
from django.core.management.base import BaseCommand

SOCIETY_ID = "<society_id>"
FLAT_ID    = "<flat_uuid>"
BASE       = "http://127.0.0.1:8000"

# (Role, Module, Method, URL, Auth-Required, Required-Role, Body/Params, Description)
# Auth-Required: "Public" | "Bearer"
# Required-Role: "None" | "Super Admin" | "Society Admin" | "Resident" | "Any"
APIS = [
    # ── AUTH / ONBOARDING — all public ───────────────────────────────────────
    ("Auth", "OTP Send",             "POST", f"{BASE}/api/accounts/otp/send/",                       "Public", "None",          '{"mobile":"9000000001"}',                          "Send OTP to mobile — dev OTP is always 123456"),
    ("Auth", "OTP Verify",           "POST", f"{BASE}/api/accounts/otp/verify/",                     "Public", "None",          '{"mobile":"9000000001","otp_code":"123456"}',       "Verify OTP before onboarding"),
    ("Auth", "Mobile OTP Login",     "POST", f"{BASE}/api/accounts/login/mobile/",                   "Public", "None",          '{"mobile":"9000000001","otp_code":"123456"}',       "Login with mobile + OTP=123456. Returns JWT + role + home_route + features"),
    ("Auth", "Email Login",          "POST", f"{BASE}/api/accounts/login/email/",                    "Public", "None",          '{"email":"superadmin@minigate.in","password":"Admin@123"}', "Login with email + password. Returns JWT + profile"),
    ("Auth", "Token Refresh",        "POST", f"{BASE}/api/accounts/token/refresh/",                  "Public", "None",          '{"refresh":"<refresh_token>"}',                    "Get new access token using refresh token"),
    ("Auth", "Countries",            "GET",  f"{BASE}/api/accounts/onboarding/countries/",           "Public", "None",          "",                                                 "List all active countries for onboarding"),
    ("Auth", "Cities",               "GET",  f"{BASE}/api/accounts/onboarding/cities/?country=1",    "Public", "None",          "?country=<id>",                                    "List cities filtered by country"),
    ("Auth", "Societies",            "GET",  f"{BASE}/api/accounts/onboarding/societies/?city=1",    "Public", "None",          "?city=<id>",                                       "List active societies in city"),
    ("Auth", "Buildings",            "GET",  f"{BASE}/api/accounts/onboarding/buildings/?society=11","Public", "None",          "?society=<id>",                                    "List buildings in society"),
    ("Auth", "Flats",                "GET",  f"{BASE}/api/accounts/onboarding/flats/?society=11",    "Public", "None",          "?society=<id> or ?building=<uuid>",                "List flats for building/society"),
    ("Auth", "Onboarding Complete",  "POST", f"{BASE}/api/accounts/onboarding/complete/",            "Public", "None",          '{"mobile","full_name","country_id","city_id","society_id","flat_number"}', "Register resident — status=PENDING until admin approves"),
    ("Auth", "Approval Status",      "GET",  f"{BASE}/api/accounts/onboarding/approval-status/",     "Public", "None",          "?mobile=9100000001",                               "Poll registration approval status (3-step progress)"),
    ("Auth", "My Home",              "GET",  f"{BASE}/api/accounts/my-home/",                        "Public", "None",          "?mobile=9100000001",                               "Get flat+building+society for resident home screen"),
    ("Auth", "My Profile",           "GET",  f"{BASE}/api/accounts/me/",                             "Bearer", "Any",           "",                                                 "Current authenticated user profile + role + permissions"),

    # ── PLATFORM ADMIN — Dashboard ───────────────────────────────────────────
    ("Platform Admin", "Dashboard Stats",      "GET",  f"{BASE}/api/platform-admin/dashboard/stats/",                          "Bearer", "Super Admin", "",                                   "Platform KPIs: total societies, users, open tickets, MRR"),
    ("Platform Admin", "Society List (KPI)",   "GET",  f"{BASE}/api/platform-admin/dashboard/societies/",                      "Bearer", "Super Admin", "?search ?status ?plan ?city ?page",  "Paginated society list with user_count + open_tickets"),

    # ── PLATFORM ADMIN — Society Management ──────────────────────────────────
    ("Platform Admin", "Society List",         "GET",  f"{BASE}/api/platform-admin/society-management/",                       "Bearer", "Super Admin", "?search ?status ?plan ?city ?page",  "Full society list with flat_count annotation"),
    ("Platform Admin", "Society Create",       "POST", f"{BASE}/api/platform-admin/society-management/",                       "Bearer", "Super Admin", '{"name","city","total_flats","plan","admin_email"}', "Create society — auto-logs platform audit"),
    ("Platform Admin", "Society Detail",       "GET",  f"{BASE}/api/platform-admin/society-management/<id>/",                  "Bearer", "Super Admin", "",                                   "Full society detail"),
    ("Platform Admin", "Society Update",       "PATCH",f"{BASE}/api/platform-admin/society-management/<id>/",                  "Bearer", "Super Admin", '{"name","plan","status",...}',        "Update society — logs audit"),
    ("Platform Admin", "Society Approve",      "POST", f"{BASE}/api/platform-admin/society-management/<id>/approve/",          "Bearer", "Super Admin", "",                                   "Approve pending society — logs platform audit"),
    ("Platform Admin", "Society Suspend",      "POST", f"{BASE}/api/platform-admin/society-management/<id>/suspend/",          "Bearer", "Super Admin", '{"reason":"..."}',                   "Suspend active society — logs platform audit"),
    ("Platform Admin", "Society Activate",     "POST", f"{BASE}/api/platform-admin/society-management/<id>/activate/",         "Bearer", "Super Admin", "",                                   "Re-activate society — logs platform audit"),
    ("Platform Admin", "Society Delete",       "DELETE",f"{BASE}/api/platform-admin/society-management/<id>/",                 "Bearer", "Super Admin", "",                                   "Delete society — logs platform audit"),

    # ── PLATFORM ADMIN — Subscription Plans ──────────────────────────────────
    ("Platform Admin", "Plans List",           "GET",  f"{BASE}/api/platform-admin/subscription-plans/",                       "Bearer", "Super Admin", "",                                   "List all subscription plans"),
    ("Platform Admin", "Plan Create",          "POST", f"{BASE}/api/platform-admin/subscription-plans/",                       "Bearer", "Super Admin", '{"name","price","billing_cycle","max_flats","features"}', "Create plan — logs audit"),
    ("Platform Admin", "Plan Detail",          "GET",  f"{BASE}/api/platform-admin/subscription-plans/<id>/",                  "Bearer", "Super Admin", "",                                   "Plan detail"),
    ("Platform Admin", "Plan Update",          "PATCH",f"{BASE}/api/platform-admin/subscription-plans/<id>/",                  "Bearer", "Super Admin", '{"name","price",...}',                "Update plan — logs 'Pro -> Enterprise' name change"),
    ("Platform Admin", "Plan Delete",          "DELETE",f"{BASE}/api/platform-admin/subscription-plans/<id>/",                 "Bearer", "Super Admin", "",                                   "Delete plan — logs audit"),

    # ── PLATFORM ADMIN — Global Users ────────────────────────────────────────
    ("Platform Admin", "Users List",           "GET",  f"{BASE}/api/platform-admin/global-users/",                             "Bearer", "Super Admin", "?search ?role ?status",               "Platform-wide user list"),
    ("Platform Admin", "User Invite",          "POST", f"{BASE}/api/platform-admin/global-users/invite/",                      "Bearer", "Super Admin", '{"email","role","society",...}',       "Invite user — logs platform audit"),
    ("Platform Admin", "User Suspend",         "POST", f"{BASE}/api/platform-admin/global-users/<id>/suspend/",                "Bearer", "Super Admin", "",                                   "Suspend user — logs audit"),
    ("Platform Admin", "User Activate",        "POST", f"{BASE}/api/platform-admin/global-users/<id>/activate/",               "Bearer", "Super Admin", "",                                   "Activate user — logs audit"),

    # ── PLATFORM ADMIN — Global Reports ──────────────────────────────────────
    ("Platform Admin", "Revenue Report",       "GET",  f"{BASE}/api/platform-admin/global-reports/revenue/",                   "Bearer", "Super Admin", "",                                   "Platform revenue breakdown by plan and month"),
    ("Platform Admin", "Growth Report",        "GET",  f"{BASE}/api/platform-admin/global-reports/growth/",                    "Bearer", "Super Admin", "",                                   "Society + user growth trend"),
    ("Platform Admin", "Occupancy Report",     "GET",  f"{BASE}/api/platform-admin/global-reports/occupancy/",                 "Bearer", "Super Admin", "",                                   "Flat occupancy across all societies"),
    ("Platform Admin", "Complaints Report",    "GET",  f"{BASE}/api/platform-admin/global-reports/complaints/",                "Bearer", "Super Admin", "",                                   "Complaint volume and resolution rate"),
    ("Platform Admin", "Visitors Report",      "GET",  f"{BASE}/api/platform-admin/global-reports/visitors/",                  "Bearer", "Super Admin", "",                                   "Visitor traffic platform-wide"),
    ("Platform Admin", "Payments Report",      "GET",  f"{BASE}/api/platform-admin/global-reports/payments/",                  "Bearer", "Super Admin", "",                                   "Maintenance payment collection report"),

    # ── PLATFORM ADMIN — Audit Logs ──────────────────────────────────────────
    ("Platform Admin", "Audit Logs",           "GET",  f"{BASE}/api/platform-admin/audit-logs/",                               "Bearer", "Super Admin", "?search ?action_type ?page",          "Platform audit trail — actor · action · target · time_ago"),
    ("Platform Admin", "Audit Logs Export",    "GET",  f"{BASE}/api/platform-admin/audit-logs/export/",                        "Bearer", "Super Admin", "",                                   "Download platform audit logs as CSV"),

    # ── SOCIETY ADMIN — Dashboard ─────────────────────────────────────────────
    ("Society Admin", "Dashboard",             "GET",  f"{BASE}/api/society-admin/dashboard/?society={SOCIETY_ID}",            "Bearer", "Society Admin", "?society=<id>",                      "KPIs: residents, pending dues, visitors inside, active complaints"),

    # ── SOCIETY ADMIN — Buildings ─────────────────────────────────────────────
    ("Society Admin", "Buildings List",        "GET",  f"{BASE}/api/society-admin/buildings/?society={SOCIETY_ID}",            "Bearer", "Society Admin", "?society",                            "List buildings with flat count"),
    ("Society Admin", "Building Create",       "POST", f"{BASE}/api/society-admin/buildings/",                                 "Bearer", "Society Admin", '{"society","name","floors","type"}',  "Add new building"),
    ("Society Admin", "Building Detail",       "GET",  f"{BASE}/api/society-admin/buildings/<id>/",                            "Bearer", "Society Admin", "",                                   "Building detail"),
    ("Society Admin", "Building Update",       "PATCH",f"{BASE}/api/society-admin/buildings/<id>/",                            "Bearer", "Society Admin", '{"name","floors"}',                  "Update building"),
    ("Society Admin", "Building Delete",       "DELETE",f"{BASE}/api/society-admin/buildings/<id>/",                           "Bearer", "Society Admin", "",                                   "Delete building"),

    # ── SOCIETY ADMIN — Flats ─────────────────────────────────────────────────
    ("Society Admin", "Flats List",            "GET",  f"{BASE}/api/society-admin/flats/?society={SOCIETY_ID}",                "Bearer", "Society Admin", "?building ?status",                   "List flats for building"),
    ("Society Admin", "Flat Create",           "POST", f"{BASE}/api/society-admin/flats/",                                     "Bearer", "Society Admin", '{"building","flat_number","type"}',   "Add flat"),
    ("Society Admin", "Flat Detail",           "GET",  f"{BASE}/api/society-admin/flats/<id>/",                                "Bearer", "Society Admin", "",                                   "Flat detail"),
    ("Society Admin", "Flat Update",           "PATCH",f"{BASE}/api/society-admin/flats/<id>/",                                "Bearer", "Society Admin", '{"status","owner_name",...}',         "Update flat"),

    # ── SOCIETY ADMIN — Residents ─────────────────────────────────────────────
    ("Society Admin", "Residents List",        "GET",  f"{BASE}/api/society-admin/residents/?society={SOCIETY_ID}",            "Bearer", "Society Admin", "?society ?status ?search",            "All residents in the society"),
    ("Society Admin", "Resident Detail",       "GET",  f"{BASE}/api/society-admin/residents/<id>/",                            "Bearer", "Society Admin", "",                                   "Resident profile"),
    ("Society Admin", "Resident Deactivate",   "POST", f"{BASE}/api/society-admin/residents/<id>/deactivate/",                 "Bearer", "Society Admin", "",                                   "Deactivate resident account"),
    ("Society Admin", "Resident Activate",     "POST", f"{BASE}/api/society-admin/residents/<id>/activate/",                   "Bearer", "Society Admin", "",                                   "Activate / approve resident account"),

    # ── SOCIETY ADMIN — Staff & Guards ────────────────────────────────────────
    ("Society Admin", "Staff List",            "GET",  f"{BASE}/api/society-admin/staff-guards/?society={SOCIETY_ID}",         "Bearer", "Society Admin", "?society ?role ?status",              "List staff members"),
    ("Society Admin", "Staff KPI",             "GET",  f"{BASE}/api/society-admin/staff-guards/kpi/?society={SOCIETY_ID}",     "Bearer", "Society Admin", "?society",                            "Staff count by role and shift"),
    ("Society Admin", "Staff Create",          "POST", f"{BASE}/api/society-admin/staff-guards/",                              "Bearer", "Society Admin", '{"society","full_name","role","shift","phone"}', "Add staff member"),
    ("Society Admin", "Staff Update",          "PATCH",f"{BASE}/api/society-admin/staff-guards/<id>/",                         "Bearer", "Society Admin", '{"status","shift","gate_assigned"}',  "Update staff"),
    ("Society Admin", "Staff Delete",          "DELETE",f"{BASE}/api/society-admin/staff-guards/<id>/",                        "Bearer", "Society Admin", "",                                   "Remove staff member"),

    # ── SOCIETY ADMIN — Vendors ───────────────────────────────────────────────
    ("Society Admin", "Vendors List",          "GET",  f"{BASE}/api/society-admin/vendors/?society={SOCIETY_ID}",              "Bearer", "Society Admin", "?society ?status ?category",          "List vendors"),
    ("Society Admin", "Vendor KPI",            "GET",  f"{BASE}/api/society-admin/vendors/kpi/?society={SOCIETY_ID}",          "Bearer", "Society Admin", "?society",                            "Vendor stats"),
    ("Society Admin", "Vendor Create",         "POST", f"{BASE}/api/society-admin/vendors/",                                   "Bearer", "Society Admin", '{"society","name","category","contact_phone","contract_start","contract_end"}', "Add vendor"),
    ("Society Admin", "Vendor Update",         "PATCH",f"{BASE}/api/society-admin/vendors/<id>/",                              "Bearer", "Society Admin", '{"status","contract_end",...}',        "Update vendor"),
    ("Society Admin", "Vendor Delete",         "DELETE",f"{BASE}/api/society-admin/vendors/<id>/",                             "Bearer", "Society Admin", "",                                   "Remove vendor"),

    # ── SOCIETY ADMIN — Visitors ──────────────────────────────────────────────
    ("Society Admin", "Visitors List",         "GET",  f"{BASE}/api/society-admin/visitors/?society={SOCIETY_ID}",             "Bearer", "Society Admin", "?society ?status ?visit_type ?flat",  "All visitors with filters"),
    ("Society Admin", "Visitor Dashboard",     "GET",  f"{BASE}/api/society-admin/visitors/dashboard/?society={SOCIETY_ID}",   "Bearer", "Society Admin", "?society",                            "KPIs: today's visitors, inside, pending, rejected"),
    ("Society Admin", "Visitor Create",        "POST", f"{BASE}/api/society-admin/visitors/",                                  "Bearer", "Society Admin", '{"society","flat","full_name","mobile","visit_type","purpose"}', "Log new visitor"),
    ("Society Admin", "Visitor Approve",       "POST", f"{BASE}/api/society-admin/visitors/<id>/approve/",                     "Bearer", "Society Admin", "{}",                                  "Approve visitor — logs society audit"),
    ("Society Admin", "Visitor Reject",        "POST", f"{BASE}/api/society-admin/visitors/<id>/reject/",                      "Bearer", "Society Admin", '{"reason":"..."}',                    "Reject visitor — logs society audit"),
    ("Society Admin", "Visitor Check-In",      "POST", f"{BASE}/api/society-admin/visitors/<id>/check-in/",                    "Bearer", "Society Admin", "",                                   "Mark visitor INSIDE — logs society audit"),
    ("Society Admin", "Visitor Check-Out",     "POST", f"{BASE}/api/society-admin/visitors/<id>/check-out/",                   "Bearer", "Society Admin", "",                                   "Mark visitor EXITED — logs society audit"),

    # ── SOCIETY ADMIN — Approvals ─────────────────────────────────────────────
    ("Society Admin", "Approvals List",        "GET",  f"{BASE}/api/society-admin/approvals/?society={SOCIETY_ID}",            "Bearer", "Society Admin", "?society ?status ?stage ?priority",   "List approval requests"),
    ("Society Admin", "Approvals KPI",         "GET",  f"{BASE}/api/society-admin/approvals/kpi/?society={SOCIETY_ID}",        "Bearer", "Society Admin", "?society",                            "Counts: pending / approved / rejected"),
    ("Society Admin", "Approval Create",       "POST", f"{BASE}/api/society-admin/approvals/",                                 "Bearer", "Society Admin", '{"society","title","category","priority","description"}', "Create approval request"),
    ("Society Admin", "Approval Approve",      "POST", f"{BASE}/api/society-admin/approvals/<id>/approve/",                    "Bearer", "Society Admin", '{"reviewer_notes":"..."}',             "Approve — logs society audit"),
    ("Society Admin", "Approval Reject",       "POST", f"{BASE}/api/society-admin/approvals/<id>/reject/",                     "Bearer", "Society Admin", '{"reason":"..."}',                    "Reject — logs society audit"),
    ("Society Admin", "Approval Progress",     "PATCH",f"{BASE}/api/society-admin/approvals/<id>/progress/",                   "Bearer", "Society Admin", '{"progress": 75}',                   "Update progress 0–100"),

    # ── SOCIETY ADMIN — Complaints ────────────────────────────────────────────
    ("Society Admin", "Complaints List",       "GET",  f"{BASE}/api/society-admin/complaints/?society={SOCIETY_ID}",           "Bearer", "Society Admin", "?society ?status ?category ?priority ?flat", "All complaints — filterable"),
    ("Society Admin", "Complaint Stats",       "GET",  f"{BASE}/api/society-admin/complaints/stats/?society={SOCIETY_ID}",     "Bearer", "Society Admin", "?society",                            "KPIs: open / in_progress / resolved_30d / high_priority"),
    ("Society Admin", "Log Complaint",         "POST", f"{BASE}/api/society-admin/complaints/log/",                            "Bearer", "Society Admin", '{"society","flat","resident","title","category","priority"}', "Log complaint on behalf of resident"),
    ("Society Admin", "Complaint Detail",      "GET",  f"{BASE}/api/society-admin/complaints/<id>/",                           "Bearer", "Society Admin", "",                                   "Full complaint detail"),
    ("Society Admin", "Complaint Assign",      "POST", f"{BASE}/api/society-admin/complaints/<id>/assign/",                    "Bearer", "Society Admin", '{"assigned_to":"<profile_id>"}',       "Assign to staff — logs audit"),
    ("Society Admin", "Complaint In-Progress", "POST", f"{BASE}/api/society-admin/complaints/<id>/in-progress/",               "Bearer", "Society Admin", "",                                   "Move Open to In Progress"),
    ("Society Admin", "Complaint Resolve",     "POST", f"{BASE}/api/society-admin/complaints/<id>/resolve/",                   "Bearer", "Society Admin", '{"resolution_notes":"..."}',           "Resolve — logs audit"),
    ("Society Admin", "Complaint Close",       "POST", f"{BASE}/api/society-admin/complaints/<id>/close/",                     "Bearer", "Society Admin", "",                                   "Close complaint"),

    # ── SOCIETY ADMIN — Payments ──────────────────────────────────────────────
    ("Society Admin", "Payments Overview",     "GET",  f"{BASE}/api/society-admin/payments/overview/?society={SOCIETY_ID}",    "Bearer", "Society Admin", "?society",                            "KPIs: collected / pending / overdue totals"),
    ("Society Admin", "Dues List",             "GET",  f"{BASE}/api/society-admin/payments/?society={SOCIETY_ID}",             "Bearer", "Society Admin", "?society ?status ?flat ?month",       "Paginated dues list"),
    ("Society Admin", "Due Create",            "POST", f"{BASE}/api/society-admin/payments/",                                  "Bearer", "Society Admin", '{"society","flat","amount","month","due_date"}', "Create maintenance due"),
    ("Society Admin", "Mark Paid",             "POST", f"{BASE}/api/society-admin/payments/<id>/mark-paid/",                    "Bearer", "Society Admin", '{"payment_mode":"cash"}',             "Mark due as paid"),
    ("Society Admin", "Mark Overdue",          "POST", f"{BASE}/api/society-admin/payments/<id>/mark-overdue/",                 "Bearer", "Society Admin", "",                                   "Mark due as overdue"),

    # ── SOCIETY ADMIN — Maintenance Expenses ─────────────────────────────────
    ("Society Admin", "Expenses List",         "GET",  f"{BASE}/api/society-admin/maintenance-expenses/?society={SOCIETY_ID}", "Bearer", "Society Admin", "?society ?category ?is_published",    "List expenses"),
    ("Society Admin", "Expense Create",        "POST", f"{BASE}/api/society-admin/maintenance-expenses/",                      "Bearer", "Society Admin", '{"society","title","category","amount","expense_date","vendor_name"}', "Record expense"),
    ("Society Admin", "Expense Detail",        "GET",  f"{BASE}/api/society-admin/maintenance-expenses/<id>/",                 "Bearer", "Society Admin", "",                                   "Expense detail"),
    ("Society Admin", "Expense Update",        "PATCH",f"{BASE}/api/society-admin/maintenance-expenses/<id>/",                 "Bearer", "Society Admin", '{"amount","notes","is_published"}',   "Update / publish to residents"),
    ("Society Admin", "Expense Delete",        "DELETE",f"{BASE}/api/society-admin/maintenance-expenses/<id>/",                "Bearer", "Society Admin", "",                                   "Delete expense"),

    # ── SOCIETY ADMIN — Fund Dashboard ───────────────────────────────────────
    ("Society Admin", "Fund Dashboard",        "GET",  f"{BASE}/api/society-admin/fund-dashboard/?society={SOCIETY_ID}",       "Bearer", "Society Admin", "?society",                            "6 KPI cards + fund usage bar (% expenses used) + latest expenses table"),

    # ── SOCIETY ADMIN — Monthly Statements ───────────────────────────────────
    ("Society Admin", "Statements List",       "GET",  f"{BASE}/api/society-admin/monthly-statements/?society={SOCIETY_ID}",   "Bearer", "Society Admin", "?society",                            "All monthly statements for society"),
    ("Society Admin", "Statement Detail",      "GET",  f"{BASE}/api/society-admin/monthly-statements/<id>/",                   "Bearer", "Society Admin", "",                                   "Statement detail with financials + proof docs"),
    ("Society Admin", "Generate Statement",    "POST", f"{BASE}/api/society-admin/monthly-statements/generate/",               "Bearer", "Society Admin", '{"society","month":"2026-05-01","opening_balance":85000}', "Generate or update statement for month"),
    ("Society Admin", "Publish Statement",     "POST", f"{BASE}/api/society-admin/monthly-statements/<id>/publish/",           "Bearer", "Society Admin", "",                                   "Publish to residents — logs society audit"),
    ("Society Admin", "Unpublish Statement",   "POST", f"{BASE}/api/society-admin/monthly-statements/<id>/unpublish/",         "Bearer", "Society Admin", "",                                   "Unpublish statement"),
    ("Society Admin", "Upload Proof PDFs",     "POST", f"{BASE}/api/society-admin/monthly-statements/<id>/upload-proof/",      "Bearer", "Society Admin", "multipart: files[] (4-5 PDFs)",       "Upload up to 10 proof documents at once"),
    ("Society Admin", "Delete Proof",          "DELETE",f"{BASE}/api/society-admin/monthly-statements/<id>/delete-proof/",     "Bearer", "Society Admin", "?doc_id=<id>",                        "Remove a specific proof document"),
    ("Society Admin", "Download PDF",          "GET",  f"{BASE}/api/society-admin/monthly-statements/<id>/download-pdf/",      "Bearer", "Society Admin", "",                                   "Stream full PDF statement (reportlab)"),
    ("Society Admin", "Export Excel",          "GET",  f"{BASE}/api/society-admin/monthly-statements/<id>/export-excel/",      "Bearer", "Society Admin", "",                                   "Download statement as .xlsx (openpyxl)"),

    # ── SOCIETY ADMIN — Analytics ─────────────────────────────────────────────
    ("Society Admin", "Analytics",             "GET",  f"{BASE}/api/society-admin/analytics/?society={SOCIETY_ID}&period=30d", "Bearer", "Society Admin", "?society ?period=7d|30d|90d|1y",      "Overview KPIs + collection trend + expense trend + complaint chart + visitor chart + occupancy"),

    # ── SOCIETY ADMIN — Notice Board ──────────────────────────────────────────
    ("Society Admin", "Notices List",          "GET",  f"{BASE}/api/society-admin/notice-board/?society={SOCIETY_ID}",         "Bearer", "Society Admin", "?society ?category ?status",          "List notices / events / fundraisers"),
    ("Society Admin", "Notice Create",         "POST", f"{BASE}/api/society-admin/notice-board/",                              "Bearer", "Society Admin", '{"society","title","description","category","audience"}', "Post notice"),
    ("Society Admin", "Notice Update",         "PATCH",f"{BASE}/api/society-admin/notice-board/<id>/",                         "Bearer", "Society Admin", '{"status","event_date",...}',          "Update notice"),
    ("Society Admin", "Notice Delete",         "DELETE",f"{BASE}/api/society-admin/notice-board/<id>/",                        "Bearer", "Society Admin", "",                                   "Delete notice"),

    # ── SOCIETY ADMIN — Security ──────────────────────────────────────────────
    ("Society Admin", "Security Dashboard",    "GET",  f"{BASE}/api/society-admin/security/dashboard/?society={SOCIETY_ID}",   "Bearer", "Society Admin", "?society",                            "Security KPIs: gates open/closed, active alerts, visitors inside"),
    ("Society Admin", "Gates List",            "GET",  f"{BASE}/api/society-admin/security/gates/?society={SOCIETY_ID}",       "Bearer", "Society Admin", "?society",                            "Gate list and status"),
    ("Society Admin", "Gate Update",           "PATCH",f"{BASE}/api/society-admin/security/gates/<id>/",                       "Bearer", "Society Admin", '{"status":"open"}',                   "Open / close gate"),
    ("Society Admin", "Security Alerts",       "GET",  f"{BASE}/api/society-admin/security/alerts/?society={SOCIETY_ID}",      "Bearer", "Society Admin", "?society ?status ?alert_type",        "Security alert list"),

    # ── SOCIETY ADMIN — Audit Logs ────────────────────────────────────────────
    ("Society Admin", "Society Audit Logs",    "GET",  f"{BASE}/api/society-admin/audit-logs/?society={SOCIETY_ID}",           "Bearer", "Society Admin", "?society ?search ?action_type ?actor_role", "Society audit trail: Admin * Priya | action | target | 15 min ago"),
    ("Society Admin", "Audit Logs Export",     "GET",  f"{BASE}/api/society-admin/audit-logs/export/?society={SOCIETY_ID}",    "Bearer", "Society Admin", "?society",                            "CSV export of society audit logs"),

    # ── RESIDENT — Dashboard ──────────────────────────────────────────────────
    ("Resident", "Dashboard",                  "GET",  f"{BASE}/api/resident/dashboard/?society={SOCIETY_ID}&flat={FLAT_ID}",  "Bearer", "Resident",      "?society ?flat",                      "Home screen: pending dues, unread notices, active complaints, visitor history"),

    # ── RESIDENT — Profile ────────────────────────────────────────────────────
    ("Resident", "Profile",                    "GET",  f"{BASE}/api/resident/profile/",                                        "Bearer", "Resident",      "",                                   "Own profile: name, flat, society, role"),
    ("Resident", "Profile Update",             "PATCH",f"{BASE}/api/resident/profile/",                                        "Bearer", "Resident",      '{"full_name","mobile","photo_url"}',  "Update own profile"),

    # ── RESIDENT — Complaints ─────────────────────────────────────────────────
    ("Resident", "My Complaints",              "GET",  f"{BASE}/api/resident/complaints/",                                     "Bearer", "Resident",      "?status ?category",                   "Own complaints list"),
    ("Resident", "Raise Complaint",            "POST", f"{BASE}/api/resident/complaints/",                                     "Bearer", "Resident",      '{"society","flat","title","category","description","priority"}', "Submit new complaint"),
    ("Resident", "Complaint Detail",           "GET",  f"{BASE}/api/resident/complaints/<id>/",                                "Bearer", "Resident",      "",                                   "Own complaint status + resolution notes"),

    # ── RESIDENT — Payments ───────────────────────────────────────────────────
    ("Resident", "My Dues",                    "GET",  f"{BASE}/api/resident/payments/",                                       "Bearer", "Resident",      "?status ?month",                      "Own maintenance dues list"),
    ("Resident", "Due Detail",                 "GET",  f"{BASE}/api/resident/payments/<id>/",                                  "Bearer", "Resident",      "",                                   "Due detail"),
    ("Resident", "Pay Due",                    "POST", f"{BASE}/api/resident/payments/<id>/pay/",                              "Bearer", "Resident",      '{"payment_mode":"online"}',            "Mark own due as paid"),

    # ── RESIDENT — Notices ────────────────────────────────────────────────────
    ("Resident", "Society Notices",            "GET",  f"{BASE}/api/resident/notices/?society={SOCIETY_ID}",                   "Bearer", "Resident",      "?society ?category",                  "View society notices and events"),
    ("Resident", "Mark Notice Read",           "POST", f"{BASE}/api/resident/notices/<id>/read/",                              "Bearer", "Resident",      "",                                   "Mark notice as read"),

    # ── RESIDENT — Visitors ───────────────────────────────────────────────────
    ("Resident", "My Visitors",                "GET",  f"{BASE}/api/resident/visitors/",                                       "Bearer", "Resident",      "?status",                             "Own visitor history"),
    ("Resident", "Invite Visitor",             "POST", f"{BASE}/api/resident/visitors/",                                       "Bearer", "Resident",      '{"society","flat","full_name","mobile","visit_type","purpose"}', "Pre-register visitor"),

    # ── RESIDENT — Maintenance Transparency ──────────────────────────────────
    ("Resident", "Published Expenses",         "GET",  f"{BASE}/api/resident/maintenance-transparency/?society={SOCIETY_ID}",  "Bearer", "Resident",      "?society ?month",                     "View published expenses + monthly statements"),

    # ── RESIDENT — SOS ────────────────────────────────────────────────────────
    ("Resident", "SOS Alert",                  "POST", f"{BASE}/api/resident/sos/",                                            "Bearer", "Resident",      '{"society","message"}',               "Send emergency SOS alert"),

    # ── SECURITY GUARD — Dashboard ────────────────────────────────────────────
    ("Security Guard", "Guard Dashboard",           "GET",  f"{BASE}/api/security-guard/dashboard/",                              "Bearer", "Security Guard", "",                                   "4 KPI cards: in_today, out_today, at_gate, alerts + shift info"),

    # ── SECURITY GUARD — Gate Entry ───────────────────────────────────────────
    ("Security Guard", "Gate Entry List",           "GET",  f"{BASE}/api/security-guard/gate-entry/",                             "Bearer", "Security Guard", "?search ?visit_type ?status ?date",  "Paginated gate entry list"),
    ("Security Guard", "Gate Entry Create",         "POST", f"{BASE}/api/security-guard/gate-entry/",                             "Bearer", "Security Guard", '{"flat","visitor_name","mobile","visit_type","purpose"}', "Log new visitor at gate"),
    ("Security Guard", "Gate Entry Summary",        "GET",  f"{BASE}/api/security-guard/gate-entry/summary/",                     "Bearer", "Security Guard", "",                                   "KPIs: today_entries, today_exits, currently_inside, pending_approval"),
    ("Security Guard", "Entry/Exit Log",            "GET",  f"{BASE}/api/security-guard/gate-entry/log/",                         "Bearer", "Security Guard", "?search ?visit_type ?status ?date",  "Today's gate movement: NAME, TYPE, IN time, OUT time, STATUS"),
    ("Security Guard", "Entry/Exit Log Export",     "GET",  f"{BASE}/api/security-guard/gate-entry/log/export/",                  "Bearer", "Security Guard", "?date=2026-05-25",                   "CSV export of entry/exit log"),
    ("Security Guard", "Gate Entry Detail",         "GET",  f"{BASE}/api/security-guard/gate-entry/<pk>/",                        "Bearer", "Security Guard", "",                                   "Single gate entry detail"),

    # ── SECURITY GUARD — Visitor Log ──────────────────────────────────────────
    ("Security Guard", "Visitor Log List",          "GET",  f"{BASE}/api/security-guard/visitor-log/",                            "Bearer", "Security Guard", "?status ?visit_type ?date ?search",  "All visitors with status filters"),
    ("Security Guard", "Visitor Approve",           "POST", f"{BASE}/api/security-guard/visitor-log/<pk>/approve/",               "Bearer", "Security Guard", "",                                   "Approve pending visitor"),
    ("Security Guard", "Visitor Reject",            "POST", f"{BASE}/api/security-guard/visitor-log/<pk>/reject/",                "Bearer", "Security Guard", '{"reason":"..."}',                   "Reject visitor with optional reason"),
    ("Security Guard", "Visitor Check-In",          "POST", f"{BASE}/api/security-guard/visitor-log/<pk>/check-in/",              "Bearer", "Security Guard", "",                                   "Mark visitor INSIDE — records checked_in_at"),
    ("Security Guard", "Visitor Check-Out",         "POST", f"{BASE}/api/security-guard/visitor-log/<pk>/check-out/",             "Bearer", "Security Guard", "",                                   "Mark visitor EXITED — records checked_out_at"),

    # ── SECURITY GUARD — Vehicle Tracking ─────────────────────────────────────
    ("Security Guard", "Vehicle List",              "GET",  f"{BASE}/api/security-guard/vehicle-tracking/",                       "Bearer", "Security Guard", "?vehicle_type ?status ?date ?search","Vehicle log list"),
    ("Security Guard", "Log Vehicle",               "POST", f"{BASE}/api/security-guard/vehicle-tracking/",                       "Bearer", "Security Guard", '{"vehicle_number","vehicle_type","direction","flat"}', "Log vehicle IN or OUT"),
    ("Security Guard", "Vehicle Summary",           "GET",  f"{BASE}/api/security-guard/vehicle-tracking/summary/",               "Bearer", "Security Guard", "",                                   "KPIs: vehicles_in_today, vehicles_out_today, currently_inside, by_type"),
    ("Security Guard", "Vehicle Detail",            "GET",  f"{BASE}/api/security-guard/vehicle-tracking/<pk>/",                  "Bearer", "Security Guard", "",                                   "Single vehicle log detail"),

    # ── SECURITY GUARD — Emergency Alerts ─────────────────────────────────────
    ("Security Guard", "Alerts List",               "GET",  f"{BASE}/api/security-guard/emergency-alerts/",                       "Bearer", "Security Guard", "?status=active|acknowledged|resolved","All emergency alerts with status filter"),
    ("Security Guard", "Create Alert",              "POST", f"{BASE}/api/security-guard/emergency-alerts/",                       "Bearer", "Security Guard", '{"alert_type","description","location"}', "Raise new emergency alert"),
    ("Security Guard", "Alert Stats",               "GET",  f"{BASE}/api/security-guard/emergency-alerts/stats/",                 "Bearer", "Security Guard", "",                                   "KPI counts: active, acknowledged, total"),
    ("Security Guard", "Alert Detail",              "GET",  f"{BASE}/api/security-guard/emergency-alerts/<pk>/",                  "Bearer", "Security Guard", "",                                   "Single alert detail"),
    ("Security Guard", "Acknowledge Alert",         "POST", f"{BASE}/api/security-guard/emergency-alerts/<pk>/acknowledge/",      "Bearer", "Security Guard", "",                                   "Move alert active → acknowledged"),
    ("Security Guard", "Resolve Alert",             "POST", f"{BASE}/api/security-guard/emergency-alerts/<pk>/resolve/",          "Bearer", "Security Guard", '{"notes":"..."}',                    "Move alert acknowledged → resolved"),

    # ── SECURITY GUARD — Shift Management ─────────────────────────────────────
    ("Security Guard", "Shift List",                "GET",  f"{BASE}/api/security-guard/shift-management/",                       "Bearer", "Security Guard", "?date ?guard",                       "All shifts with date/guard filters"),
    ("Security Guard", "Today Shift",               "GET",  f"{BASE}/api/security-guard/shift-management/today/",                 "Bearer", "Security Guard", "",                                   "Guard's own shift for today (null if none scheduled)"),

    # ── SECURITY GUARD — Visitor Entry ────────────────────────────────────────
    ("Security Guard", "Register Visitor",          "POST", f"{BASE}/api/security-guard/visitor-entry/",                          "Bearer", "Security Guard", '{"flat","visitor_name","mobile","visit_type","purpose"}', "Register new visitor at gate"),
    ("Security Guard", "Search Visitor",            "GET",  f"{BASE}/api/security-guard/visitor-entry/search/",                   "Bearer", "Security Guard", "?mobile=9800000001 or ?name=Kavya",  "Search pre-registered visitors / guest passes"),

    # ── SECURITY GUARD — QR / Passcode ────────────────────────────────────────
    ("Security Guard", "Verify QR/Passcode",        "POST", f"{BASE}/api/security-guard/qr-passcode/verify/",                     "Bearer", "Security Guard", '{"pass_code":"MG-PASS-9XK2"}',       "Verify guest pass — returns visitor details if valid"),
    ("Security Guard", "Check-In via QR",           "POST", f"{BASE}/api/security-guard/qr-passcode/checkin/",                    "Bearer", "Security Guard", '{"pass_code":"MG-PASS-9XK2"}',       "Check in visitor via QR scan or passcode"),

    # ── SECURITY GUARD — Delivery Verify ──────────────────────────────────────
    ("Security Guard", "Delivery List",             "GET",  f"{BASE}/api/security-guard/delivery-verify/",                        "Bearer", "Security Guard", "?status ?flat ?date ?search",        "All deliveries with filters"),
    ("Security Guard", "Pending Deliveries",        "GET",  f"{BASE}/api/security-guard/delivery-verify/pending/",                "Bearer", "Security Guard", "",                                   "Deliveries awaiting guard action"),
    ("Security Guard", "Deliveries At Gate",        "GET",  f"{BASE}/api/security-guard/delivery-verify/at-gate/",                "Bearer", "Security Guard", "",                                   "Deliveries currently at gate"),
    ("Security Guard", "Delivery Summary",          "GET",  f"{BASE}/api/security-guard/delivery-verify/summary/",                "Bearer", "Security Guard", "",                                   "KPIs: today_total, pending, collected, rejected, returned"),
    ("Security Guard", "Create Delivery",           "POST", f"{BASE}/api/security-guard/delivery-verify/",                        "Bearer", "Security Guard", '{"flat","delivery_type","sender_name","delivery_person","mobile","description"}', "Log delivery arrival at gate"),
    ("Security Guard", "Delivery Detail",           "GET",  f"{BASE}/api/security-guard/delivery-verify/<pk>/",                   "Bearer", "Security Guard", "",                                   "Single delivery with OTP status"),
    ("Security Guard", "Approve Delivery",          "POST", f"{BASE}/api/security-guard/delivery-verify/<pk>/approve/",           "Bearer", "Security Guard", "",                                   "Approve — triggers OTP send to resident"),
    ("Security Guard", "Generate OTP",              "POST", f"{BASE}/api/security-guard/delivery-verify/<pk>/generate-otp/",      "Bearer", "Security Guard", "",                                   "Resend OTP to resident (dev OTP always 123456)"),
    ("Security Guard", "Verify OTP",                "POST", f"{BASE}/api/security-guard/delivery-verify/<pk>/verify-otp/",        "Bearer", "Security Guard", '{"otp":"123456"}',                   "Resident gives OTP to guard — marks collected on success"),
    ("Security Guard", "Reject Delivery",           "POST", f"{BASE}/api/security-guard/delivery-verify/<pk>/reject/",            "Bearer", "Security Guard", '{"reason":"..."}',                   "Reject delivery"),
    ("Security Guard", "Mark At Gate",              "POST", f"{BASE}/api/security-guard/delivery-verify/<pk>/at-gate/",           "Bearer", "Security Guard", "",                                   "Mark delivery as at_gate"),
    ("Security Guard", "Mark Collected",            "POST", f"{BASE}/api/security-guard/delivery-verify/<pk>/collected/",         "Bearer", "Security Guard", "",                                   "Mark delivery as collected"),
    ("Security Guard", "Return Delivery",           "POST", f"{BASE}/api/security-guard/delivery-verify/<pk>/return/",            "Bearer", "Security Guard", "",                                   "Mark delivery returned to sender"),

    # ── SECURITY GUARD — Approved Visitors ────────────────────────────────────
    ("Security Guard", "Approved Visitors List",    "GET",  f"{BASE}/api/security-guard/approved-visitors/",                      "Bearer", "Security Guard", "?visit_type ?search",                "Pre-approved passes + resident-approved visitors"),
    ("Security Guard", "Approved Visitor Check-In", "POST", f"{BASE}/api/security-guard/approved-visitors/checkin/",              "Bearer", "Security Guard", '{"visitor_id":3,"source":"visitor"}', "Check in an approved visitor or guest pass"),
    ("Security Guard", "Approved Visitors Export",  "GET",  f"{BASE}/api/security-guard/approved-visitors/export/",               "Bearer", "Security Guard", "",                                   "CSV export of approved visitors"),

    # ── SECURITY GUARD — Contact Resident ─────────────────────────────────────
    ("Security Guard", "Contact Resident List",     "GET",  f"{BASE}/api/security-guard/contact-resident/",                       "Bearer", "Security Guard", "?search=flat/building/name/mobile",  "All flats in society with resident mobile numbers"),
    ("Security Guard", "Contact Flat Detail",       "GET",  f"{BASE}/api/security-guard/contact-resident/{FLAT_ID}/",             "Bearer", "Security Guard", "",                                   "Contact card for one flat — all residents + mobiles"),
]


ROLE_COLORS = {
    "Auth":            ("1A1A2E", "FFFFFF"),
    "Platform Admin":  ("1B4332", "FFFFFF"),
    "Society Admin":   ("1C3A5E", "FFFFFF"),
    "Resident":        ("6B2D8B", "FFFFFF"),
    "Security Guard":  ("1A5276", "FFFFFF"),
}

METHOD_COLORS = {
    "GET":    "27AE60",
    "POST":   "2980B9",
    "PATCH":  "D35400",
    "PUT":    "F39C12",
    "DELETE": "C0392B",
}

AUTH_COLORS = {
    "Public": ("27AE60", "FFFFFF"),
    "Bearer": ("E67E22", "FFFFFF"),
}

ROLE_REQ_COLORS = {
    "None":           ("AAAAAA", "FFFFFF"),
    "Any":            ("2980B9", "FFFFFF"),
    "Super Admin":    ("1B4332", "FFFFFF"),
    "Society Admin":  ("1C3A5E", "FFFFFF"),
    "Resident":       ("6B2D8B", "FFFFFF"),
    "Security Guard": ("1A5276", "FFFFFF"),
}


class Command(BaseCommand):
    help = "Generate MiniGate_API_Report.xlsx with full API + auth documentation."

    def handle(self, *args, **options):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.stderr.write("openpyxl not installed. Run: pip install openpyxl")
            return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        ws_sum  = wb.create_sheet("Summary")
        ws_auth = wb.create_sheet("Auth & Security")
        ws_all  = wb.create_sheet("All APIs")

        for role in ["Auth", "Platform Admin", "Society Admin", "Resident", "Security Guard"]:
            ws = wb.create_sheet(role)
            role_apis = [r for r in APIS if r[0] == role]
            self._build_role_sheet(ws, role, role_apis, Font, PatternFill, Alignment, Border, Side, get_column_letter)

        ws_seed = wb.create_sheet("Seed Data")

        self._build_summary(ws_sum,  Font, PatternFill, Alignment, Border, Side, get_column_letter)
        self._build_auth_sheet(ws_auth, Font, PatternFill, Alignment, Border, Side, get_column_letter)
        self._build_all_apis(ws_all, Font, PatternFill, Alignment, Border, Side, get_column_letter)
        self._build_seed_sheet(ws_seed, Font, PatternFill, Alignment, Border, Side, get_column_letter)

        out = "MiniGate_API_Report.xlsx"
        wb.save(out)
        self.stdout.write(self.style.SUCCESS(f"Excel saved: {out}"))
        self.stdout.write(f"  Total APIs: {len(APIS)} | Sheets: Summary, Auth & Security, All APIs, Auth, Platform Admin, Society Admin, Resident, Security Guard, Seed Data")

    # ── helpers ───────────────────────────────────────────────────────────────

    def _border(self, Side, Border):
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)

    def _hdr(self, ws, cols, row, bg, fg, Font, PatternFill, Alignment, Border, Side):
        fill = PatternFill("solid", fgColor=bg)
        font = Font(color=fg, bold=True, name="Calibri", size=10)
        brd  = self._border(Side, Border)
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill, cell.font, cell.border = fill, font, brd
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 22

    def _row(self, ws, vals, row, Font, PatternFill, Alignment, Border, Side, even=False):
        fill = PatternFill("solid", fgColor="F2F6FA" if even else "FFFFFF")
        brd  = self._border(Side, Border)
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill, cell.border = fill, brd
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 18

    def _badge(self, ws, row, col, text, bg, fg, Font, PatternFill, Border, Side, Alignment):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(color=fg, bold=True, name="Calibri", size=9)
        cell.border = self._border(Side, Border)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Summary sheet ─────────────────────────────────────────────────────────

    def _build_summary(self, ws, Font, PatternFill, Alignment, Border, Side, get_column_letter):
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 40

        ws.merge_cells("A1:C1")
        ws["A1"].value = "MiniGate Backend - API Documentation"
        ws["A1"].font  = Font(name="Calibri", size=15, bold=True, color="1C3A5E")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:C2")
        ws["A2"].value = f"Generated: {datetime.datetime.now().strftime('%d %b %Y  %H:%M')}   |   Total APIs: {len(APIS)}"
        ws["A2"].font  = Font(name="Calibri", size=9, color="888888", italic=True)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 16

        self._hdr(ws, ["Role", "# APIs", "Auth Required"], 4, "1C3A5E", "FFFFFF",
                  Font, PatternFill, Alignment, Border, Side)

        roles = ["Auth", "Platform Admin", "Society Admin", "Resident", "Security Guard"]
        auth_label = ["Public (13) + Bearer (1)", "Bearer — Super Admin only",
                      "Bearer — Society Admin only", "Bearer — Resident only",
                      "Bearer — Security Guard (also: Society Admin, Super Admin)"]
        for i, (role, alabel) in enumerate(zip(roles, auth_label)):
            count = len([r for r in APIS if r[0] == role])
            bg, fg = ROLE_COLORS[role]
            r = 5 + i
            for col, val in [(1, role), (2, count), (3, alabel)]:
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill   = PatternFill("solid", fgColor=bg)
                cell.font   = Font(color=fg, name="Calibri", size=10, bold=(col < 3))
                cell.border = self._border(Side, Border)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[r].height = 22

        r = 5 + len(roles)
        for col, val in [(1, "TOTAL"), (2, len(APIS)), (3, "")]:
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill   = PatternFill("solid", fgColor="2C2C2C")
            cell.font   = Font(color="FFFFFF", name="Calibri", size=11, bold=True)
            cell.border = self._border(Side, Border)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 24

        r += 2
        ws.merge_cells(f"A{r}:C{r}")
        ws[f"A{r}"].value = "Test Accounts  (password: Admin@123  |  OTP: 123456)"
        ws[f"A{r}"].font  = Font(name="Calibri", size=11, bold=True, color="1C3A5E")
        ws[f"A{r}"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 20

        self._hdr(ws, ["Role", "Email / Mobile", "home_route"], r + 1,
                  "2C2C2C", "FFFFFF", Font, PatternFill, Alignment, Border, Side)

        creds = [
            ("Super Admin",        "superadmin@minigate.in  |  9000000001",  "platform_admin_dashboard"),
            ("Society Admin",      "societyadmin@minigate.in  |  9000000002","society_admin_dashboard"),
            ("Resident (A-101)",   "resident1@minigate.in  |  9100000001",   "resident_dashboard"),
            ("Resident (B-101)",   "resident2@minigate.in  |  9100000002",   "resident_dashboard"),
            ("Resident (B-201)",   "resident3@minigate.in  |  9100000003",   "resident_dashboard"),
        ]
        for j, (role, cred, route) in enumerate(creds, r + 2):
            for col, val in [(1, role), (2, cred), (3, route)]:
                cell = ws.cell(row=j, column=col, value=val)
                cell.font = Font(name="Calibri", size=9, bold=(col == 1))
                cell.border = self._border(Side, Border)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[j].height = 16

    # ── Auth & Security sheet ─────────────────────────────────────────────────

    def _build_auth_sheet(self, ws, Font, PatternFill, Alignment, Border, Side, get_column_letter):
        ws.sheet_view.showGridLines = False

        # Title
        ws.merge_cells("A1:E1")
        ws["A1"].value = "MiniGate — Authentication & Authorization Rules"
        ws["A1"].font  = Font(name="Calibri", size=13, bold=True, color="1C3A5E")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # How login works
        ws.merge_cells("A2:E2")
        ws["A2"].value = "LOGIN FLOW:  POST /api/accounts/login/mobile/  with  { mobile, otp_code: '123456' }  ->  Returns: tokens + home_route + features[]"
        ws["A2"].font  = Font(name="Calibri", size=10, bold=True, color="C0392B")
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 18

        ws.merge_cells("A3:E3")
        ws["A3"].value = "HEADER:  Authorization: Bearer <access_token>   (required for all non-public endpoints)"
        ws["A3"].font  = Font(name="Calibri", size=10, color="555555")
        ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[3].height = 16

        # Role access matrix
        ws.row_dimensions[5].height = 20
        self._hdr(ws, ["Endpoint Group", "No Token", "Resident Token", "Society Admin Token", "Super Admin Token"],
                  5, "2C2C2C", "FFFFFF", Font, PatternFill, Alignment, Border, Side)

        matrix = [
            ("/api/accounts/login/mobile/  (and all public endpoints)", "200 OK", "200 OK", "200 OK", "200 OK"),
            ("/api/resident/*  (personal dues, complaints, notices)",    "401",     "200 OK",  "403 Denied", "200 OK"),
            ("/api/society-admin/*  (manage society, residents, etc.)",  "401",     "403 Denied","200 OK",  "200 OK"),
            ("/api/platform-admin/*  (platform-wide management)",        "401",     "403 Denied","403 Denied","200 OK"),
        ]
        STATUS_COLORS = {
            "200 OK":     ("27AE60", "FFFFFF"),
            "401":        ("888888", "FFFFFF"),
            "403 Denied": ("C0392B", "FFFFFF"),
        }
        for i, (group, *statuses) in enumerate(matrix, 6):
            ws.cell(row=i, column=1, value=group).font = Font(name="Calibri", size=9, bold=True)
            ws.cell(row=i, column=1).border = self._border(Side, Border)
            ws.cell(row=i, column=1).alignment = Alignment(wrap_text=True, vertical="center")
            for j, st in enumerate(statuses, 2):
                bg, fg = STATUS_COLORS.get(st, ("FFFFFF", "000000"))
                self._badge(ws, i, j, st, bg, fg, Font, PatternFill, Border, Side, Alignment)
            ws.row_dimensions[i].height = 20

        # Column widths
        ws.column_dimensions["A"].width = 52
        for col in ["B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 18

        # Permission classes section
        r = 12
        ws.merge_cells(f"A{r}:E{r}")
        ws[f"A{r}"].value = "Permission Classes  (apps/common/permissions.py)"
        ws[f"A{r}"].font  = Font(name="Calibri", size=11, bold=True, color="1C3A5E")
        ws[f"A{r}"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 22

        self._hdr(ws, ["Class", "Slug Check", "Applied To", "Fallback for is_superuser", "Imported As"],
                  r + 1, "1C3A5E", "FFFFFF", Font, PatternFill, Alignment, Border, Side)

        perms = [
            ("IsSuperAdmin",             "super-admin",  "All /api/platform-admin/* views",    "YES - always passes", "from apps.common.permissions import IsSuperAdmin"),
            ("IsSocietyAdmin",           "society-admin","All /api/society-admin/* views",      "YES (via IsSuperAdmin base)", "from apps.common.permissions import IsSocietyAdmin"),
            ("IsResident",               "resident",     "All /api/resident/* views",           "YES - super admin can debug", "from apps.common.permissions import IsResident"),
            ("IsSocietyAdminOrSuperAdmin","society-admin OR super-admin","Shared endpoints",    "YES",  "from apps.common.permissions import IsSocietyAdminOrSuperAdmin"),
            ("IsSecurityGuard",          "security-guard","All /api/security-guard/* views",    "YES",  "from apps.common.permissions import IsSecurityGuard"),
            ("IsAccountant",             "accountant",   "All /api/accountant/* views",         "YES",  "from apps.common.permissions import IsAccountant"),
            ("AllowAny (DRF built-in)",  "—",            "13 public /api/accounts/* endpoints", "—",   "from rest_framework.permissions import AllowAny"),
        ]
        for i, row_data in enumerate(perms, r + 2):
            self._row(ws, list(row_data), i, Font, PatternFill, Alignment, Border, Side, i % 2 == 0)

        # Error responses
        r2 = r + 2 + len(perms) + 2
        ws.merge_cells(f"A{r2}:E{r2}")
        ws[f"A{r2}"].value = "Standard Error Responses"
        ws[f"A{r2}"].font  = Font(name="Calibri", size=11, bold=True, color="1C3A5E")
        ws[f"A{r2}"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[r2].height = 22

        self._hdr(ws, ["HTTP Status", "When", "Response Body", "", ""], r2 + 1,
                  "2C2C2C", "FFFFFF", Font, PatternFill, Alignment, Border, Side)

        errors = [
            ("401 Unauthorized", "No token / expired token",    '{"detail": "Authentication credentials were not provided."}'),
            ("401 Unauthorized", "Invalid / malformed token",   '{"detail": "Given token not valid for any token type"}'),
            ("403 Forbidden",    "Valid token but wrong role",  '{"detail": "Access denied. [Role] role required."}'),
            ("400 Bad Request",  "Wrong OTP",                   '{"success": false, "message": "Incorrect OTP. Use 123456."}'),
            ("403 Forbidden",    "Account pending approval",    '{"success": false, "message": "Your account is pending approval..."}'),
            ("403 Forbidden",    "Account deactivated",         '{"success": false, "message": "Your account has been deactivated..."}'),
            ("404 Not Found",    "Mobile not registered",       '{"success": false, "message": "No account found for this mobile number..."}'),
        ]
        for i, (code, when, body) in enumerate(errors, r2 + 2):
            self._row(ws, [code, when, body, "", ""], i, Font, PatternFill, Alignment, Border, Side, i % 2 == 0)

    # ── All APIs sheet ────────────────────────────────────────────────────────

    def _build_all_apis(self, ws, Font, PatternFill, Alignment, Border, Side, get_column_letter):
        ws.sheet_view.showGridLines = False
        headers = ["#", "Role", "Module", "Method", "Auth", "Required Role", "URL", "Body / Params", "Description"]
        widths  = [5, 14, 20, 8, 8, 14, 58, 38, 46]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        self._hdr(ws, headers, 1, "1C3A5E", "FFFFFF", Font, PatternFill, Alignment, Border, Side)

        for idx, api in enumerate(APIS, 1):
            role, module, method, url, auth_req, req_role, body, desc = api
            r   = idx + 1
            bg, fg = ROLE_COLORS.get(role, ("FFFFFF", "000000"))
            even = idx % 2 == 0
            self._row(ws, [idx, role, module, "", "", "", url, body, desc],
                      r, Font, PatternFill, Alignment, Border, Side, even)

            # method badge
            self._badge(ws, r, 4, method,
                        METHOD_COLORS.get(method, "888888"), "FFFFFF",
                        Font, PatternFill, Border, Side, Alignment)

            # auth badge (Public green / Bearer orange)
            abg, afg = AUTH_COLORS.get(auth_req, ("888888", "FFFFFF"))
            self._badge(ws, r, 5, auth_req, abg, afg, Font, PatternFill, Border, Side, Alignment)

            # required role badge
            rbg, rfg = ROLE_REQ_COLORS.get(req_role, ("888888", "FFFFFF"))
            self._badge(ws, r, 6, req_role, rbg, rfg, Font, PatternFill, Border, Side, Alignment)

            # role column tinted
            cell = ws.cell(row=r, column=2)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(color=fg, name="Calibri", size=9, bold=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Per-role sheet ────────────────────────────────────────────────────────

    def _build_role_sheet(self, ws, role, role_apis, Font, PatternFill, Alignment, Border, Side, get_column_letter):
        bg, fg = ROLE_COLORS.get(role, ("1C3A5E", "FFFFFF"))
        ws.sheet_view.showGridLines = False
        headers = ["#", "Module", "Method", "Auth", "Required Role", "URL", "Body / Params", "Description"]
        widths  = [5, 22, 9, 8, 14, 62, 40, 50]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        title = ws["A1"]
        title.value = f"{role}  -  {len(role_apis)} APIs"
        title.fill  = PatternFill("solid", fgColor=bg)
        title.font  = Font(color=fg, name="Calibri", size=13, bold=True)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        self._hdr(ws, headers, 2, "2C2C2C", "FFFFFF", Font, PatternFill, Alignment, Border, Side)

        for idx, api in enumerate(role_apis, 1):
            _, module, method, url, auth_req, req_role, body, desc = api
            r = idx + 2
            self._row(ws, [idx, module, "", "", "", url, body, desc],
                      r, Font, PatternFill, Alignment, Border, Side, idx % 2 == 0)
            self._badge(ws, r, 3, method,
                        METHOD_COLORS.get(method, "888888"), "FFFFFF",
                        Font, PatternFill, Border, Side, Alignment)
            abg, afg = AUTH_COLORS.get(auth_req, ("888888", "FFFFFF"))
            self._badge(ws, r, 4, auth_req, abg, afg, Font, PatternFill, Border, Side, Alignment)
            rbg, rfg = ROLE_REQ_COLORS.get(req_role, ("888888", "FFFFFF"))
            self._badge(ws, r, 5, req_role, rbg, rfg, Font, PatternFill, Border, Side, Alignment)

        ws.freeze_panes = "A3"

    # ── Seed Data sheet ───────────────────────────────────────────────────────

    def _build_seed_sheet(self, ws, Font, PatternFill, Alignment, Border, Side, get_column_letter):
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 62
        ws.column_dimensions["C"].width = 12

        self._hdr(ws, ["Category", "Seeded Data", "Count"], 1,
                  "1C3A5E", "FFFFFF", Font, PatternFill, Alignment, Border, Side)

        rows = [
            ("Society",              "Greenwood Heights, Bengaluru  |  Pro plan  |  Active",              "1"),
            ("Buildings",            "Tower A (A-101, A-102, A-201)   Tower B (B-101, B-201, B-202)",       "2"),
            ("Flats",                "6 flats across Tower A and Tower B",                                 "6"),
            ("Residents",            "resident1/2/3@minigate.in  |  mobile: 9100000001/2/3",               "3"),
            ("Staff Members",        "7: Security Guards x3, Housekeeping x2, Maintenance, Gardener",      "7"),
            ("Gates",                "Gate 1 Main (Open), Gate 2 Service (Open), Gate 3 Exit (Closed)",    "3"),
            ("Vendors",              "AquaPure, GreenScape, FixIt Plumbing, CleanPro, LiftTech",           "5"),
            ("Notices",              "Notice, Fundraiser, Maintenance event",                              "3"),
            ("Maintenance Dues",     "Current month: 2 Paid, 2 Pending, 2 Overdue",                        "6"),
            ("Visitors",             "INSIDE x2, PENDING x2, APPROVED x1, EXITED x1",                     "6"),
            ("Approval Requests",    "Move-in, Vehicle Reg, Pet Reg, Renovation, Domestic Help",           "6"),
            ("Complaints",           "Water leak (IN_PROGRESS), Lift (OPEN), Noise, Parking, Garden (RESOLVED)", "5"),
            ("Security Alerts",      "Unauthorized vehicle, Tailgating, Camera offline",                   "3"),
            ("Maintenance Expenses", "Security salary, Water tanker, Lift AMC, Garden, Electricity, Plumbing, Housekeeping", "7"),
            ("Monthly Statement",    "May 2026 - Published - collected 78,500 / expenses 1,16,200 / balance -37,700", "1"),
            ("Platform Tickets",     "Login OTP issue, Subscription failed, Feature requests x2, Account lock, iOS PDF bug", "6"),
            ("Platform Payments",    "Subscription (paid), Setup fee, Analytics addon, Upcoming renewal",  "4"),
            ("Platform Audit Logs",  "create/approve/suspend society, invite user, update plan",            "5"),
            ("Society Audit Logs",   "approve visitor, resolve complaint, assign complaint, publish statement, check-in", "7"),
            ("Module Permissions",   "Super Admin: 17 modules ALL CRUD  |  Society Admin: 14 modules ALL CRUD  |  Resident: 4 modules VIEW/CREATE", "3 roles"),
        ]
        for i, (cat, detail, cnt) in enumerate(rows, 2):
            self._row(ws, [cat, detail, cnt], i, Font, PatternFill, Alignment, Border, Side, i % 2 == 0)

        r = len(rows) + 4
        ws.merge_cells(f"A{r}:C{r}")
        ws[f"A{r}"].value = "python manage.py seed_data --clear --settings=config.settings.development"
        ws[f"A{r}"].font  = Font(name="Courier New", size=10, bold=True, color="C0392B")
        ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 20

        r2 = r + 1
        ws.merge_cells(f"A{r2}:C{r2}")
        ws[f"A{r2}"].value = "python manage.py generate_api_excel --settings=config.settings.development"
        ws[f"A{r2}"].font  = Font(name="Courier New", size=10, bold=True, color="1B4332")
        ws[f"A{r2}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r2].height = 20
